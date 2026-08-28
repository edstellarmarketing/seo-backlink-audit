"""
DA / PA / Spam Score providers.

WHY THERE IS NO "SCRAPE THE FREE CHECKER" PROVIDER
--------------------------------------------------
The popular free bulk checkers (tools.guestpostlinks.net, dapachecker.org,
dapacheckerpro.com) are protected by Cloudflare Turnstile and device
fingerprinting, and dapachecker.org states its free tier uses a CAPTCHA.
Automating those submissions means defeating bot detection: it breaks the
moment they rotate the challenge, and it is not something this project does.

So there are two supported routes, both legitimate:

  provider: import       You run the bulk checker yourself in your browser
                         (that is what it is for) and drop the result in
                         input/metrics/. The audit writes ready-to-paste
                         batches to output/da_pa_queue.txt to make this quick.
                         Free. ~2 minutes of clicking per 100 domains.

  provider: moz          Official Moz Links API. Real DA, PA and Spam Score.
  provider: dataforseo   DataForSEO Backlinks API. Cheap, pay-per-call.
  provider: rapidapi     A RapidAPI DA/PA endpoint - note that dapacheckerpro
                         sells its own "Bulk Moz DA PA SS Checker" there, so
                         this is the sanctioned way to get that same data.

Every provider result is cached on disk by registered domain, so you never
pay for or re-paste a domain twice within `metrics.cache_days`.
"""

import csv
import json
import os
import re
import time

import requests

from seo_audit.analysis import domains as dom

# --------------------------------------------------------------------------
# Cache
# --------------------------------------------------------------------------
EMPTY = {
    "da": None, "pa": None, "spam_score": None,
    "backlinks": None, "quality_backlinks": None, "referring_domains": None,
    "domain_age": None, "source": "", "fetched_at": None,
}


class MetricsCache:
    def __init__(self, path: str, cache_days: int = 30):
        self.path = path
        self.ttl = cache_days * 86400
        self.data = {}
        self._load()

    def _load(self):
        if os.path.exists(self.path):
            try:
                with open(self.path, encoding="utf-8") as f:
                    self.data = json.load(f)
            except (json.JSONDecodeError, OSError):
                self.data = {}

    def save(self):
        os.makedirs(os.path.dirname(self.path) or ".", exist_ok=True)
        tmp = self.path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=1)
        os.replace(tmp, self.path)

    def get(self, domain: str):
        rec = self.data.get(domain.lower())
        if not rec:
            return None
        if self.ttl and rec.get("fetched_at"):
            if time.time() - rec["fetched_at"] > self.ttl:
                return None
        return rec

    def put(self, domain: str, rec: dict):
        rec = {**EMPTY, **rec}
        rec["fetched_at"] = rec.get("fetched_at") or time.time()
        self.data[domain.lower()] = rec

    def stats(self):
        return len(self.data)


# --------------------------------------------------------------------------
# Column detection for imported checker exports
# --------------------------------------------------------------------------
_COL_PATTERNS = [
    ("da", re.compile(r"^\s*(da|domain\s*authority|domain_authority|moz\s*da|dom\s*auth)\s*%?\s*$", re.I)),
    ("pa", re.compile(r"^\s*(pa|page\s*authority|page_authority|moz\s*pa)\s*%?\s*$", re.I)),
    ("spam_score", re.compile(r"^\s*(ss|spam\s*score|spam_score|spamscore|spam)\s*%?\s*$", re.I)),
    ("backlinks", re.compile(r"^\s*(tb|total\s*backlinks|backlinks|total_backlinks|inbound\s*links)\s*$", re.I)),
    ("quality_backlinks", re.compile(r"^\s*(qb|quality\s*backlinks|dofollow|do\s*follow|quality_backlinks)\s*$", re.I)),
    ("referring_domains", re.compile(r"^\s*(rd|referring\s*domains|linking\s*domains|ref\s*domains)\s*$", re.I)),
    ("domain_age", re.compile(r"^\s*(age|domain\s*age|domain_age)\s*$", re.I)),
]
_URL_PATTERN = re.compile(r"^\s*(url|urls|website|websites|domain|domains|web\s*page|link|site|address)\s*$", re.I)

_NUM = re.compile(r"-?\d+(?:\.\d+)?")


def _num(v):
    """Pull the first number out of a cell. '42%' -> 42.0, '1,234' -> 1234.0"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s or s.upper() in ("N/A", "NA", "-", "--", "NULL", "NONE", "ERROR"):
        return None
    m = _NUM.search(s)
    return float(m.group()) if m else None


def _looks_like_domain(s: str) -> bool:
    s = str(s or "").strip()
    if not s or " " in s.strip():
        return bool(re.match(r"^https?://\S+$", s))
    return bool(re.match(r"^(https?://)?[\w.-]+\.[a-z]{2,}(/.*)?$", s, re.I))


def parse_metric_rows(rows: list) -> dict:
    """
    Turn rows-of-cells from any DA/PA checker export into
    {registered_domain: {da, pa, spam_score, ...}}.

    Handles guestpostlinks (DA/PA/SS/TB/QB), dapachecker, dapacheckerpro,
    and generic pasted tables. Finds the header row wherever it sits.
    """
    if not rows:
        return {}

    # Locate the header row: the first row where >=2 cells match our patterns
    header_idx, colmap, url_col = None, {}, None
    for idx, row in enumerate(rows[:25]):
        cells = [str(c or "").strip() for c in row]
        found, ucol = {}, None
        for j, c in enumerate(cells):
            if _URL_PATTERN.match(c) and ucol is None:
                ucol = j
                continue
            for key, pat in _COL_PATTERNS:
                if pat.match(c) and key not in found:
                    found[key] = j
        if len(found) >= 2:
            header_idx, colmap, url_col = idx, found, ucol
            break

    out = {}

    if header_idx is not None:
        # If no explicit URL header, use the first column that holds domains
        if url_col is None:
            for j in range(len(rows[header_idx])):
                sample = [r[j] for r in rows[header_idx + 1: header_idx + 6]
                          if len(r) > j]
                if sample and sum(_looks_like_domain(s) for s in sample) >= max(1, len(sample) // 2):
                    url_col = j
                    break
        if url_col is None:
            url_col = 0

        for row in rows[header_idx + 1:]:
            if len(row) <= url_col:
                continue
            raw = str(row[url_col] or "").strip()
            reg = dom.registered_domain(raw)
            if not reg or not _looks_like_domain(raw):
                continue
            rec = {}
            for key, j in colmap.items():
                if len(row) > j:
                    rec[key] = _num(row[j])
            if any(v is not None for v in rec.values()):
                out[reg] = rec
        return out

    # ---- No header found: assume "domain  DA  PA  SS ..." positional -----
    for row in rows:
        cells = [str(c or "").strip() for c in row if str(c or "").strip() != ""]
        if len(cells) < 2 or not _looks_like_domain(cells[0]):
            continue
        reg = dom.registered_domain(cells[0])
        if not reg:
            continue
        nums = [_num(c) for c in cells[1:]]
        nums = [n for n in nums if n is not None]
        if not nums:
            continue
        rec = {}
        for key, val in zip(("da", "pa", "spam_score", "backlinks", "quality_backlinks"), nums):
            rec[key] = val
        out[reg] = rec
    return out


def _rows_from_file(path: str) -> list:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            for r in ws.iter_rows(values_only=True):
                rows.append(list(r))
        wb.close()
        return rows
    if ext in (".csv", ".tsv"):
        with open(path, newline="", encoding="utf-8-sig", errors="ignore") as f:
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            return [list(r) for r in csv.reader(f, dialect)]
    # .txt / anything else: split on tabs or runs of 2+ spaces
    with open(path, encoding="utf-8-sig", errors="ignore") as f:
        rows = []
        for line in f:
            line = line.rstrip("\n")
            if not line.strip():
                continue
            parts = re.split(r"\t+|\s{2,}", line.strip())
            if len(parts) == 1:
                parts = line.strip().split()
            rows.append(parts)
        return rows


# --------------------------------------------------------------------------
# Providers
# --------------------------------------------------------------------------
class BaseProvider:
    name = "none"

    def fetch(self, domains: list) -> dict:
        return {}


class NoneProvider(BaseProvider):
    name = "none"


class ImportProvider(BaseProvider):
    """Reads DA/PA/SS from checker exports you drop in input/metrics/."""

    name = "import"

    def __init__(self, metrics_dir: str):
        self.dir = metrics_dir

    def fetch(self, domains: list) -> dict:
        if not os.path.isdir(self.dir):
            return {}
        merged = {}

        def is_data_file(name: str) -> bool:
            """
            Skip documentation and Office lock files. This matters: the
            instructions file shipped in input/metrics/ contains EXAMPLE tables,
            and without this guard they were parsed as real metrics and injected
            invented DA values into the audit.
            """
            base = name.lower()
            if base.startswith("~$") or base.startswith("."):
                return False
            if base.split(".")[0] in ("readme", "read-me", "instructions", "notes", "help"):
                return False
            return os.path.splitext(base)[1] in (".csv", ".tsv", ".txt", ".xlsx", ".xlsm")

        files = sorted(
            (os.path.join(self.dir, f) for f in os.listdir(self.dir) if is_data_file(f)),
            key=os.path.getmtime,
        )
        for path in files:
            try:
                rows = _rows_from_file(path)
                found = parse_metric_rows(rows)
            except Exception as e:                     # noqa: BLE001
                print(f"  ! could not parse {os.path.basename(path)}: {type(e).__name__}: {e}")
                continue
            if found:
                print(f"  + {os.path.basename(path)}: {len(found)} domains with metrics")
            for k, v in found.items():
                merged[k] = {**merged.get(k, {}), **v,
                             "source": f"import:{os.path.basename(path)}"}
        return merged


class MozProvider(BaseProvider):
    """Official Moz Links API v2 - real DA, PA, Spam Score."""

    name = "moz"
    ENDPOINT = "https://lsapi.seomoz.com/v2/url_metrics"

    def __init__(self, access_id: str, secret: str, timeout: int = 40):
        self.auth = (access_id, secret)
        self.timeout = timeout

    def fetch(self, domains: list) -> dict:
        out = {}
        for i in range(0, len(domains), 50):        # Moz allows 50 per call
            batch = domains[i:i + 50]
            try:
                r = requests.post(
                    self.ENDPOINT, auth=self.auth, timeout=self.timeout,
                    json={"targets": batch},
                )
                if r.status_code == 401:
                    print("  ! Moz: 401 Unauthorized - check MOZ_ACCESS_ID / MOZ_SECRET")
                    return out
                r.raise_for_status()
                results = r.json().get("results", [])
            except (requests.RequestException, ValueError) as e:
                print(f"  ! Moz batch {i // 50 + 1} failed: {e}")
                continue
            for item in results:
                target = item.get("page") or item.get("subdomain") or item.get("root_domain") or ""
                reg = dom.registered_domain(target)
                if not reg:
                    continue
                out[reg] = {
                    "da": item.get("domain_authority"),
                    "pa": item.get("page_authority"),
                    "spam_score": item.get("spam_score"),
                    "backlinks": item.get("external_pages_to_root_domain") or item.get("external_pages"),
                    "referring_domains": item.get("root_domains_to_root_domain"),
                    "source": "moz-api",
                }
            print(f"  + Moz: {len(results)} results (batch {i // 50 + 1})")
        return out


class DataForSEOProvider(BaseProvider):
    """
    DataForSEO Backlinks API.

    Note honestly: DataForSEO returns its OWN rank (0-1000), not Moz DA. We
    rescale rank/10 -> a 0-100 'DA-equivalent'. Their backlink_spam_score IS a
    0-100 spam score. Column headers in the report say 'DA (DFS rank)' so you
    are never misled into thinking it is literally Moz DA.
    """

    name = "dataforseo"
    ENDPOINT = "https://api.dataforseo.com/v3/backlinks/bulk_ranks/live"
    SPAM_ENDPOINT = "https://api.dataforseo.com/v3/backlinks/bulk_spam_score/live"
    SUMMARY_ENDPOINT = "https://api.dataforseo.com/v3/backlinks/bulk_backlinks/live"

    def __init__(self, login: str, password: str, timeout: int = 60):
        self.auth = (login, password)
        self.timeout = timeout

    def _call(self, endpoint, targets):
        try:
            r = requests.post(endpoint, auth=self.auth, timeout=self.timeout,
                              json=[{"targets": targets}])
            if r.status_code == 401:
                print("  ! DataForSEO: 401 - check DATAFORSEO_LOGIN / DATAFORSEO_PASSWORD")
                return []
            r.raise_for_status()
            tasks = r.json().get("tasks") or []
            items = []
            for t in tasks:
                for res in (t.get("result") or []):
                    items.extend(res.get("items") or [])
            return items
        except (requests.RequestException, ValueError) as e:
            print(f"  ! DataForSEO call failed ({endpoint.rsplit('/', 2)[-2]}): {e}")
            return []

    def fetch(self, domains: list) -> dict:
        out = {}
        for i in range(0, len(domains), 100):
            batch = domains[i:i + 100]
            for item in self._call(self.ENDPOINT, batch):
                reg = dom.registered_domain(item.get("target", ""))
                if reg:
                    rank = item.get("rank")
                    out.setdefault(reg, {})["da"] = round(rank / 10, 1) if rank is not None else None
                    out[reg]["source"] = "dataforseo"
            for item in self._call(self.SPAM_ENDPOINT, batch):
                reg = dom.registered_domain(item.get("target", ""))
                if reg:
                    out.setdefault(reg, {})["spam_score"] = item.get("spam_score")
                    out[reg]["source"] = "dataforseo"
            for item in self._call(self.SUMMARY_ENDPOINT, batch):
                reg = dom.registered_domain(item.get("target", ""))
                if reg:
                    out.setdefault(reg, {})["backlinks"] = item.get("backlinks")
                    out[reg]["referring_domains"] = item.get("referring_domains")
                    out[reg]["source"] = "dataforseo"
            print(f"  + DataForSEO: batch {i // 100 + 1} -> {len(out)} domains so far")
        return out


class RapidAPIProvider(BaseProvider):
    """
    Generic RapidAPI DA/PA endpoint (e.g. dapacheckerpro's own
    'Bulk Moz DA PA SS Checker').

    RapidAPI endpoints differ, so this walks the JSON response and picks up
    whatever da / pa / spam-score keys it finds. Set metrics.rapidapi_path in
    config.yaml if your endpoint is not the default.
    """

    name = "rapidapi"

    def __init__(self, key: str, host: str, path: str = "/", method: str = "POST",
                 batch_size: int = 20, timeout: int = 60):
        self.key, self.host = key, host
        self.path = path if path.startswith("/") else "/" + path
        self.method = (method or "POST").upper()
        self.batch = max(1, batch_size)
        self.timeout = timeout

    @staticmethod
    def _pick(d: dict, *names):
        low = {str(k).lower().replace(" ", "").replace("_", ""): v for k, v in d.items()}
        for n in names:
            key = n.lower().replace("_", "")
            if key in low:
                return _num(low[key])
        return None

    def _walk(self, obj, out):
        """Find dicts anywhere in the response that carry a domain + metrics."""
        if isinstance(obj, dict):
            target = None
            for k in ("url", "domain", "target", "website", "site", "page"):
                for kk in obj:
                    if str(kk).lower() == k and obj[kk]:
                        target = str(obj[kk])
                        break
                if target:
                    break
            da = self._pick(obj, "da", "domainauthority", "domain_authority", "domainAuthority")
            pa = self._pick(obj, "pa", "pageauthority", "page_authority", "pageAuthority")
            ss = self._pick(obj, "ss", "spamscore", "spam_score", "spamScore")
            if target and any(v is not None for v in (da, pa, ss)):
                reg = dom.registered_domain(target)
                if reg:
                    out[reg] = {
                        "da": da, "pa": pa, "spam_score": ss,
                        "backlinks": self._pick(obj, "tb", "backlinks", "totalbacklinks"),
                        "quality_backlinks": self._pick(obj, "qb", "qualitybacklinks", "dofollow"),
                        "source": "rapidapi",
                    }
            for v in obj.values():
                self._walk(v, out)
        elif isinstance(obj, list):
            for v in obj:
                self._walk(v, out)

    def fetch(self, domains: list) -> dict:
        out = {}
        url = f"https://{self.host}{self.path}"
        headers = {
            "X-RapidAPI-Key": self.key,
            "X-RapidAPI-Host": self.host,
            "Content-Type": "application/json",
        }
        for i in range(0, len(domains), self.batch):
            batch = domains[i:i + self.batch]
            try:
                if self.method == "GET":
                    r = requests.get(url, headers=headers, timeout=self.timeout,
                                     params={"url": ",".join(batch), "domains": ",".join(batch)})
                else:
                    r = requests.post(url, headers=headers, timeout=self.timeout,
                                      json={"urls": batch, "domains": batch, "url": ",".join(batch)})
                if r.status_code in (401, 403):
                    print(f"  ! RapidAPI: {r.status_code} - check RAPIDAPI_KEY / host / subscription")
                    return out
                r.raise_for_status()
                self._walk(r.json(), out)
            except (requests.RequestException, ValueError) as e:
                print(f"  ! RapidAPI batch {i // self.batch + 1} failed: {e}")
                continue
            print(f"  + RapidAPI: batch {i // self.batch + 1} -> {len(out)} domains so far")
        return out


# --------------------------------------------------------------------------
def build_provider(cfg: dict, project_root: str) -> BaseProvider:
    m = cfg.get("metrics", {}) or {}
    name = str(m.get("provider", "none")).strip().lower()
    creds = m.get("credentials", {}) or {}

    def cred(key, env):
        return str(os.environ.get(env) or creds.get(key) or "").strip()

    if name in ("", "none", "off", "skip"):
        return NoneProvider()

    if name in ("import", "browser", "manual", "paste"):
        return ImportProvider(os.path.join(project_root, cfg.get("input_dir", "input"), "metrics"))

    if name == "moz":
        aid, sec = cred("moz_access_id", "MOZ_ACCESS_ID"), cred("moz_secret", "MOZ_SECRET")
        if not (aid and sec):
            print("  ! metrics.provider=moz but no credentials found -> falling back to import/")
            return ImportProvider(os.path.join(project_root, cfg.get("input_dir", "input"), "metrics"))
        return MozProvider(aid, sec)

    if name == "dataforseo":
        lg, pw = cred("dataforseo_login", "DATAFORSEO_LOGIN"), cred("dataforseo_password", "DATAFORSEO_PASSWORD")
        if not (lg and pw):
            print("  ! metrics.provider=dataforseo but no credentials found -> falling back to import/")
            return ImportProvider(os.path.join(project_root, cfg.get("input_dir", "input"), "metrics"))
        return DataForSEOProvider(lg, pw)

    if name == "rapidapi":
        key, host = cred("rapidapi_key", "RAPIDAPI_KEY"), cred("rapidapi_host", "RAPIDAPI_HOST")
        if not (key and host):
            print("  ! metrics.provider=rapidapi but no credentials found -> falling back to import/")
            return ImportProvider(os.path.join(project_root, cfg.get("input_dir", "input"), "metrics"))
        return RapidAPIProvider(
            key, host,
            path=str(m.get("rapidapi_path", "/")),
            method=str(m.get("rapidapi_method", "POST")),
            batch_size=int(m.get("rapidapi_batch", 20)),
        )

    print(f"  ! unknown metrics.provider '{name}' -> no metrics")
    return NoneProvider()


def collect(domains: list, provider: BaseProvider, cache: MetricsCache,
            allow_missing: bool = True) -> dict:
    """
    Return {domain: metrics} for every domain, using the cache first and only
    asking the provider for what is genuinely missing.
    """
    result, missing = {}, []
    for d in domains:
        rec = cache.get(d)
        if rec:
            result[d] = rec
        else:
            missing.append(d)

    if missing and not isinstance(provider, NoneProvider):
        print(f"\nFetching DA/PA for {len(missing)} uncached domain(s) via '{provider.name}'...")
        try:
            fetched = provider.fetch(missing)
        except Exception as e:                          # noqa: BLE001
            print(f"  ! provider '{provider.name}' raised {type(e).__name__}: {e}")
            fetched = {}
        for d, rec in fetched.items():
            rec.setdefault("source", provider.name)
            cache.put(d, rec)
            result[d] = cache.get(d) or rec
        cache.save()
        still = [d for d in missing if d not in fetched]
        if still:
            print(f"  - {len(still)} domain(s) still without metrics"
                  f"{' (scored on on-page signals only)' if allow_missing else ''}")

    for d in domains:
        result.setdefault(d, dict(EMPTY))
    return result


def write_queue(domains: list, path: str, batch_size: int = 100):
    """
    Write domains that still need DA/PA into ready-to-paste batches.
    This is what makes the free route quick: open the checker, paste a batch.
    """
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("# Domains still missing DA / PA / Spam Score\n")
        f.write("# Paste one BATCH at a time into a bulk DA-PA checker, e.g.\n")
        f.write("#   https://tools.guestpostlinks.net/bulk-da-pa-checker-tool/\n")
        f.write("#   https://www.dapachecker.org/\n")
        f.write("#   https://dapacheckerpro.com/\n")
        f.write("# Then save/copy the result table into  input/metrics/  and re-run.\n")
        f.write("# Any of .csv .xlsx .txt works - columns are auto-detected.\n\n")
        for i in range(0, len(domains), batch_size):
            batch = domains[i:i + batch_size]
            f.write(f"### BATCH {i // batch_size + 1}  ({len(batch)} domains)\n")
            f.write("\n".join(batch))
            f.write("\n\n")
    return (len(domains) + batch_size - 1) // batch_size
