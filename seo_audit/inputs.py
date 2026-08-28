"""
Reading backlink lists.

Column names are auto-detected so exports from Ahrefs, Semrush and Search
Console work untouched. A plain .txt of one URL per line works too.
"""

import csv
import os
import sys

from seo_audit.net import fetch

URL_KEYS = ("url", "urls", "link", "links", "backlink", "backlinks", "backlink url",
            "source url", "source", "page", "referring page", "referring_page",
            "referring url", "website", "domain", "site", "address")
TARGET_KEYS = ("target", "target url", "destination", "link to", "your url",
               "target page", "landing page", "linked url")
ANCHOR_KEYS = ("anchor", "anchor text", "anchortext", "keyword", "anchor_text")
NOTE_KEYS = ("notes", "note", "comment", "comments", "remark", "source note")

# Every column name we recognise, for header detection.
ALL_KEYS = set(URL_KEYS) | set(TARGET_KEYS) | set(ANCHOR_KEYS) | set(NOTE_KEYS)


# ---------------------------------------------------------------------------
# Input
# ---------------------------------------------------------------------------
def _match_key(fieldnames, candidates):
    for cand in candidates:
        for fn in fieldnames or []:
            if fn and fn.strip().lower() == cand:
                return fn
    for cand in candidates:                     # loose contains-match
        for fn in fieldnames or []:
            if fn and cand in fn.strip().lower():
                return fn
    return None


def read_input_file(path: str) -> list:
    """Read one .csv/.tsv/.xlsx/.txt into [{url,target,anchor,notes}]."""
    ext = os.path.splitext(path)[1].lower()
    items = []

    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            print(f"  ! openpyxl missing, cannot read {path}")
            return []
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        if not rows:
            return []
        header = [str(c or "").strip() for c in rows[0]]
        uk = _match_key(header, URL_KEYS)
        if uk is None:
            header, start = [f"col{i}" for i in range(len(rows[0]))], 0
            uk = "col0"
        else:
            start = 1
        idx = {h: i for i, h in enumerate(header)}
        tk, ak, nk = (_match_key(header, TARGET_KEYS), _match_key(header, ANCHOR_KEYS),
                      _match_key(header, NOTE_KEYS))
        for row in rows[start:]:
            get = lambda k: (str(row[idx[k]]).strip() if k and k in idx and len(row) > idx[k] and row[idx[k]] is not None else "")
            u = get(uk)
            if u and u.lower() not in ("none", "nan"):
                items.append({"url": u, "target": get(tk), "anchor": get(ak), "notes": get(nk)})
        return items

    if ext in (".csv", ".tsv"):
        with open(path, newline="", encoding="utf-8-sig", errors="ignore") as f:
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            # csv.Sniffer().has_header() is a heuristic and it gets this wrong
            # on files whose columns are all short lowercase words -- it decided
            # "url,target,anchor,notes" was data, and the audit dutifully tried
            # to fetch https://url. So: if the first cell is literally one of
            # the column names we recognise, it IS a header, whatever the
            # sniffer thinks.
            first_row = next(csv.reader([sample.splitlines()[0]], dialect), []) \
                if sample.strip() else []
            looks_like_header = bool(first_row) and any(
                str(c).strip().lower() in ALL_KEYS for c in first_row)

            has_header = looks_like_header
            if not has_header:
                try:
                    has_header = csv.Sniffer().has_header(sample)
                except csv.Error:
                    pass
            if has_header:
                rd = csv.DictReader(f, dialect=dialect)
                uk = _match_key(rd.fieldnames, URL_KEYS) or (rd.fieldnames or [None])[0]
                tk = _match_key(rd.fieldnames, TARGET_KEYS)
                ak = _match_key(rd.fieldnames, ANCHOR_KEYS)
                nk = _match_key(rd.fieldnames, NOTE_KEYS)
                for row in rd:
                    u = (row.get(uk) or "").strip()
                    if u:
                        items.append({
                            "url": u,
                            "target": (row.get(tk) or "").strip() if tk else "",
                            "anchor": (row.get(ak) or "").strip() if ak else "",
                            "notes": (row.get(nk) or "").strip() if nk else "",
                        })
            else:
                for row in csv.reader(f, dialect):
                    if row and row[0].strip():
                        if row[0].strip().lower() in ALL_KEYS:
                            continue                 # a header after all
                        items.append({"url": row[0].strip(),
                                      "target": row[1].strip() if len(row) > 1 else "",
                                      "anchor": "", "notes": ""})
        return items

    # plain text: one URL per line
    with open(path, encoding="utf-8-sig", errors="ignore") as f:
        for line in f:
            u = line.strip()
            if u and not u.startswith("#"):
                items.append({"url": u.split()[0], "target": "", "anchor": "", "notes": ""})
    return items


def gather_input(cfg: dict, explicit: str | None, urls: list | None) -> list:
    items = []
    if urls:
        items += [{"url": u, "target": "", "anchor": "", "notes": "from --urls"} for u in urls]
    if explicit:
        if not os.path.exists(explicit):
            sys.exit(f"ERROR: input file not found: {explicit}")
        items += read_input_file(explicit)
    elif not urls:
        in_dir = os.path.join(ROOT, cfg.get("input_dir", "input"))
        if not os.path.isdir(in_dir):
            sys.exit(f"ERROR: input folder not found: {in_dir}")
        found = [f for f in sorted(os.listdir(in_dir))
                 if os.path.splitext(f)[1].lower() in (".csv", ".tsv", ".txt", ".xlsx", ".xlsm")
                 and not f.startswith("~$")]
        if not found:
            sys.exit(f"ERROR: no .csv/.xlsx/.txt files in {in_dir}\n"
                     f"       Drop your backlink list there, or pass --urls a.com b.com")
        for f in found:
            got = read_input_file(os.path.join(in_dir, f))
            print(f"  + {f}: {len(got)} rows")
            items += got

    # de-duplicate on normalized URL, keep the first occurrence
    seen, unique = set(), []
    for it in items:
        u = fetch.normalize_url(it["url"])
        if not u or u in seen:
            continue
        seen.add(u)
        it["url"] = u
        unique.append(it)
    return unique
