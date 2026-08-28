"""
Report writers: Excel workbook, CSV, HTML dashboard, disavow file,
outreach list, anchor-text report and link-network report.
"""

import csv
import html as _html
import json
import os
from collections import Counter

from seo_audit.scoring import anchors as anchors_mod

# Column order for the flat exports. The useful stuff comes first -- nobody
# scrolls fifty columns to find out whether their link is still there.
COLUMNS = [
    ("url", "URL"),
    ("registered", "Domain"),
    ("verdict", "Verdict"),
    ("score", "Score"),
    ("tier", "Tier"),
    ("tier_label", "Trust tier"),
    ("da", "DA"),
    ("pa", "PA"),
    ("spam_score", "Spam Score %"),
    ("status_code", "HTTP"),
    ("status_verdict", "Status"),
    ("link_found", "Link live?"),
    ("is_followed", "Followed?"),
    ("link_placement", "Placement"),
    ("link_is_sitewide", "Sitewide?"),
    ("link_rel", "rel="),
    ("anchor_texts", "Anchor text"),
    ("is_noindex", "noindex?"),
    ("action", "What to do"),
    ("issues", "Issues found"),
    ("wins", "Positives"),
    ("gate_stage", "Reached stage"),
    ("gate_reason", "Why it stopped"),
    ("network_id", "Network cluster"),
    ("network_size", "Cluster size"),
    ("network_signals", "Cluster signals"),
    ("net_org", "Hosted by"),
    ("net_range", "IP range"),
    ("shared_host", "Shared host?"),
    ("word_count", "Words"),
    ("outbound_links", "Outbound links"),
    ("relevance_score", "Relevance %"),
    ("relevance_matched", "Relevance matched"),
    ("relevance_missing", "Relevance missing"),
    ("sitewide_ratio", "Sitewide ratio"),
    ("sitewide_sampled", "Pages sampled"),
    ("sitewide_note", "Sitewide finding"),
    ("lang", "Language"),
    ("lang_mismatch", "Lang mismatch?"),
    ("page_title", "Page title"),
    ("generator", "CMS / generator"),
    ("final_url", "Final URL"),
    ("redirect_hops", "Hops"),
    ("redirect_target_domain", "Redirects to"),
    ("redirect_chain", "Redirect chain"),
    ("content_spam_categories", "Content spam"),
    ("content_spam_keywords", "Spam keywords"),
    ("url_spam_categories", "URL spam"),
    ("parked_markers", "Parked markers"),
    ("paid_link_markers", "Paid-link markers"),
    ("link_directory_markers", "Directory markers"),
    ("home_status_code", "Home HTTP"),
    ("home_title", "Home page title"),
    ("home_content_spam_categories", "Home page spam"),
    ("home_outbound_links", "Home outbound"),
    ("home_parked_markers", "Home parked?"),
    ("free_subdomain", "Free subdomain?"),
    ("spam_tld", "Risky TLD?"),
    ("suffix", "TLD"),
    ("authority_group", "Authority group"),
    ("in_master", "On master list?"),
    ("master_status", "Master status"),
    ("master_match", "Matched by"),
    ("master_da", "Master DA"),
    ("master_spam_score", "Master SS"),
    ("master_first_seen", "First recorded"),
    ("master_disagrees", "Sheet disagrees?"),
    ("ip", "IP"),
    ("scheme_used", "Reached via"),
    ("tls_invalid", "Bad TLS cert?"),
    ("https_unavailable", "No HTTPS?"),
    ("server", "Server"),
    ("backlinks", "Total backlinks"),
    ("referring_domains", "Ref. domains"),
    ("metrics_source", "Metrics from"),
    ("archive_url", "Wayback snapshot"),
    ("archive_date", "Snapshot date"),
    ("archive_link_found", "Link in snapshot?"),
    ("archive_anchor", "Old anchor text"),
    ("archive_note", "Archive finding"),
    ("rendered", "Browser re-checked?"),
    ("render_note", "Browser finding"),
    ("root_note", "Root-domain check"),
    ("response_ms", "Response ms"),
    ("error", "Error"),
    ("notes", "Your notes"),
]

VERDICT_ORDER = ["TOXIC", "DEAD", "LINK_LOST", "BLOCKED", "LOW_VALUE", "REVIEW", "GOOD"]

VERDICT_COLORS = {          # (excel ARGB fill, html hex)
    "GOOD":      ("FFC6EFCE", "#16a34a"),
    "REVIEW":    ("FFFFEB9C", "#ca8a04"),
    "LOW_VALUE": ("FFE2E8F0", "#64748b"),
    "LINK_LOST": ("FFFFE0B2", "#ea580c"),
    "BLOCKED":   ("FFE0E0E0", "#6b7280"),
    "DEAD":      ("FFF8CBAD", "#b45309"),
    "TOXIC":     ("FFFFC7CE", "#dc2626"),
}

# HTTP status buckets used by the report's status filter.
STATUS_GROUPS = [
    ("200", "200 Live", "#16a34a"),
    ("3xx", "3xx Redirect", "#ca8a04"),
    ("404", "404 / 410 Dead", "#b45309"),
    ("5xx", "5xx Server error", "#a16207"),
    ("blocked", "401 / 403 / 429", "#6b7280"),
    ("dns", "DNS - domain gone", "#dc2626"),
    ("tls", "Broken TLS cert", "#9333ea"),
    ("error", "Other error", "#78716c"),
]
TIER_LABELS = {"A": "A - Institutional", "B": "B - High authority",
               "C": "C - Non-profit TLD", "D": "D - Unproven"}


def status_group(row: dict) -> str:
    sv = row.get("status_verdict", "")
    if sv == "LIVE":
        return "200"
    if sv.startswith("REDIRECT"):
        return "3xx"
    if sv in ("DEAD", "GONE"):
        return "404"
    if sv == "SERVER_ERROR":
        return "5xx"
    if sv in ("BLOCKED", "RATE_LIMITED"):
        return "blocked"
    if sv == "DNS_ERROR":
        return "dns"
    if sv == "SSL_ERROR":
        return "tls"
    return "error"


_BAD_SHEET = str.maketrans({c: "-" for c in "[]:*?/\\"})


def safe_sheet_name(title: str) -> str:
    """
    Excel rejects [ ] : * ? / \ in sheet names and caps them at 31 chars.
    Status labels like "401 / 403 / 429" hit both limits, so sanitise centrally
    rather than remembering to do it at each call site.
    """
    name = (title or "Sheet").translate(_BAD_SHEET).strip() or "Sheet"
    return name[:31]


def _val(row, key):
    v = row.get(key, "")
    if isinstance(v, bool):
        return "YES" if v else "NO"
    if v is None:
        return ""
    return v


def summarize(rows: list, cfg: dict | None = None) -> dict:
    cfg = cfg or {}
    v = Counter(r.get("verdict", "?") for r in rows)
    s = Counter(r.get("status_verdict", "?") for r in rows)
    t = Counter(r.get("tier", "?") for r in rows)
    g = Counter(status_group(r) for r in rows)
    scores = [r.get("score", 0) for r in rows]
    checked = [r for r in rows if r.get("target_checked")]
    das = [float(r["da"]) for r in rows if r.get("da") not in (None, "")]
    return {
        "total": len(rows),
        "verdicts": dict(v),
        "statuses": dict(s),
        "status_groups": dict(g),
        "tiers": dict(t),
        "avg_score": round(sum(scores) / len(scores), 1) if scores else 0,
        "avg_da": round(sum(das) / len(das), 1) if das else None,
        "with_metrics": len(das),
        "gate_passed": sum(1 for r in rows if r.get("gate_passed")),
        "stages": dict(Counter(r.get("gate_stage", "?") for r in rows)),
        "home_checked": sum(1 for r in rows if r.get("home_checked")),
        "home_spam": sum(1 for r in rows if r.get("home_content_spam")),
        "links_checked": len(checked),
        "links_live": sum(1 for r in checked if r.get("link_found")),
        "links_followed": sum(1 for r in checked if r.get("is_followed")),
        "links_sitewide": sum(1 for r in rows if r.get("link_is_sitewide")),
        "links_incontent": sum(1 for r in rows if r.get("link_placement") == "in-content"),
        "noindex": sum(1 for r in rows if r.get("is_noindex")),
        "lang_mismatch": sum(1 for r in rows if r.get("lang_mismatch")),
        "tls_invalid": sum(1 for r in rows if r.get("tls_invalid")),
        "no_https": sum(1 for r in rows if r.get("https_unavailable")),
        "in_network": sum(1 for r in rows if r.get("network_id")),
        "truly_sitewide": sum(1 for r in rows if r.get("sitewide_sampled", 0) >= 2
                              and r.get("sitewide_ratio", 0) >= 1.0),
        "js_recovered": sum(1 for r in rows if r.get("rendered") and r.get("link_found")),
        "archive_recovered": sum(1 for r in rows if r.get("archive_link_found")),
        "shared_hosted": sum(1 for r in rows if r.get("shared_host")),
        "in_master": sum(1 for r in rows if r.get("in_master")),
        "master_spammy": sum(1 for r in rows if r.get("in_master")
                             and "spam" in (r.get("master_status") or "").lower()),
        "master_clean": sum(1 for r in rows if r.get("in_master")
                            and "no issue" in (r.get("master_status") or "").lower()),
        "brand_new": sum(1 for r in rows if not r.get("in_master")),
        "master_disagrees": sum(1 for r in rows if master_disagreement(r)),
        "clean_now_bad": sum(1 for r in rows
                             if master_disagreement(r) == "clean_now_bad"),
        "spammy_now_clean": sum(1 for r in rows
                                if master_disagreement(r) == "spammy_now_clean"),
        "http_200": g.get("200", 0),
        "http_3xx": g.get("3xx", 0),
        "http_404": g.get("404", 0),
        "http_5xx": g.get("5xx", 0),
        "http_blocked": g.get("blocked", 0) + g.get("tls", 0),
        "http_dns": g.get("dns", 0),
        "http_ssl": g.get("tls", 0),
        "http_err": g.get("error", 0) + g.get("dns", 0),
    }


# ---------------------------------------------------------------------------
def write_csv(rows: list, path: str):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([label for _, label in COLUMNS])
        for r in rows:
            w.writerow([_val(r, k) for k, _ in COLUMNS])


def write_disavow(rows: list, path: str) -> int:
    """
    Google disavow file. ONLY includes TOXIC rows -- never LOW_VALUE, never
    LINK_LOST, never DEAD. Disavowing links that are merely useless is a
    well-known way to hurt your own rankings.
    """
    from seo_audit.analysis import domains as dom
    entries, clusters = set(), {}
    for r in rows:
        if r.get("verdict") != "TOXIC":
            continue
        if r.get("network_id"):
            clusters.setdefault(r["network_id"], r.get("network_signals", ""))
        # For free-subdomain hosts, disavow the exact host -- disavowing
        # 'zapto.org' would blanket a whole dynamic-DNS provider.
        if r.get("free_subdomain"):
            host = r.get("host") or dom.host_of(r.get("url", ""))
            if host:
                entries.add(f"domain:{host}")
        else:
            reg = r.get("registered")
            if reg:
                entries.add(f"domain:{reg}")
    with open(path, "w", encoding="utf-8") as f:
        f.write("# Google disavow file - generated by SEO Backlink Audit\n")
        f.write("# Upload: https://search.google.com/search-console/disavow-links\n#\n")
        f.write("# READ THIS FIRST. Only TOXIC rows are here - links with real spam\n")
        f.write("# signals. Links that are merely low-value (nofollow, noindex, thin)\n")
        f.write("# are deliberately EXCLUDED: disavowing harmless links can cost you\n")
        f.write("# rankings. Sanity-check every line before you upload.\n#\n")
        if clusters:
            f.write("# Some of these belong to detected link networks:\n")
            for cid, sig in sorted(clusters.items()):
                f.write(f"#   {cid}: {sig}\n")
            f.write("#\n")
        for e in sorted(entries):
            f.write(e + "\n")
    return len(entries)


def write_outreach(rows: list, path: str) -> int:
    """Links worth an email: page alive but link gone, nofollow, or sitewide-only."""
    targets = []
    for r in rows:
        if r.get("verdict") == "LINK_LOST":
            targets.append((r, "Link removed from a live page", "Ask them to restore the link"))
        elif r.get("link_found") and r.get("is_nofollow") and r.get("score", 0) >= 55:
            targets.append((r, "Link is nofollow", "Ask them to remove rel=nofollow"))
        elif r.get("link_is_sitewide") and r.get("score", 0) >= 55:
            targets.append((r, "Link is sitewide boilerplate (footer/sidebar)",
                            "Ask for one in-content link instead of a sitewide one"))
    targets.sort(key=lambda t: -(float(t[0].get("da") or 0)))
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Priority", "Domain", "URL", "DA", "Problem", "Ask"])
        for i, (r, prob, ask) in enumerate(targets, 1):
            w.writerow([i, r.get("registered", ""), r.get("url", ""),
                        r.get("da") or "", prob, ask])
    return len(targets)


def write_anchor_report(rows: list, path: str, cfg: dict) -> dict:
    rep = anchors_mod.analyse(rows, cfg)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Anchor text", "Count", "Share %", "Type"])
        total = max(1, rep["total_anchors"])
        for text, n in rep["top_anchors"]:
            w.writerow([text, n, f"{100*n/total:.1f}",
                        anchors_mod.classify_anchor(text, rep["brand_terms"])])
        w.writerow([])
        w.writerow(["Anchor type", "Count", "Share %"])
        for kind, n in sorted(rep["kinds"].items(), key=lambda kv: -kv[1]):
            w.writerow([kind, n, f"{100*n/total:.1f}"])
        if rep["warnings"]:
            w.writerow([])
            w.writerow(["Warnings"])
            for wa in rep["warnings"]:
                w.writerow([wa])
    return rep


def write_network_report(rows: list, path: str, net: dict) -> int:
    clusters = net.get("clusters", [])
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Cluster", "Size", "Shared signals", "Domains"])
        for c in clusters:
            w.writerow([c["id"], c["size"], "; ".join(c["signals"]), ", ".join(c["members"])])
        if not clusters:
            w.writerow(["(none detected)", 0, "", ""])
    return len(clusters)


# ---------------------------------------------------------------------------
def write_xlsx(rows: list, path: str, summary: dict, anchor_rep: dict | None = None,
               net: dict | None = None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        print("  ! openpyxl not installed - skipping .xlsx")
        return

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="FF1F3A5F")
    header_font = Font(bold=True, color="FFFFFFFF", size=11)

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws.append(["SEO Backlink Audit - Summary"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([])

    def block(title, pairs):
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        for k, v in pairs:
            ws.append([k, v])
        ws.append([])

    block("Overall", [
        ("Backlinks audited", summary["total"]),
        ("Average score (0-100)", summary["avg_score"]),
        ("Average DA", summary["avg_da"] if summary["avg_da"] is not None else "no DA data"),
        ("Domains with DA/PA data", f"{summary['with_metrics']} of {summary['total']}"),
    ])
    block("Verdicts", [(v, summary["verdicts"].get(v, 0)) for v in VERDICT_ORDER])
    block("HTTP status", [(label, summary["status_groups"].get(key, 0))
                          for key, label, _ in STATUS_GROUPS])
    block("Trust tiers", [(TIER_LABELS[k], summary["tiers"].get(k, 0)) for k in "ABCD"])
    block("Staged pipeline", [
        ("Stopped at stage 1 (not live)", summary["stages"].get("1-live", 0)),
        ("Stopped at stage 2 (page content)", summary["stages"].get("2-page-content", 0)),
        ("Stopped at stage 3 (home page)", summary["stages"].get("3-homepage", 0)),
        ("Reached stage 4 (DA/PA worth checking)", summary["stages"].get("4-metrics", 0)),
        ("Home pages scanned", summary.get("home_checked", 0)),
        ("Home pages found spammy", summary.get("home_spam", 0)),
    ])
    block("Link verification", [
        ("Backlinks with a target to verify", summary["links_checked"]),
        ("Link found on page", summary["links_live"]),
        ("Link found AND followed", summary["links_followed"]),
        ("In-content links", summary.get("links_incontent", 0)),
        ("Sitewide boilerplate links", summary.get("links_sitewide", 0)),
        ("Pages set to noindex", summary["noindex"]),
    ])
    block("Technical flags", [
        ("Invalid TLS certificate", summary.get("tls_invalid", 0)),
        ("HTTPS unavailable (http only)", summary.get("no_https", 0)),
        ("Language mismatch", summary.get("lang_mismatch", 0)),
        ("Domains inside a detected link network", summary.get("in_network", 0)),
        ("On a large shared host (IP discounted)", summary.get("shared_hosted", 0)),
        ("Confirmed sitewide by sampling", summary.get("truly_sitewide", 0)),
        ("Only visible after JavaScript ran", summary.get("js_recovered", 0)),
        ("Dead pages recovered from the archive", summary.get("archive_recovered", 0)),
    ])
    if anchor_rep:
        block("Anchor text", [("Total anchors seen", anchor_rep["total_anchors"]),
                              ("Distinct anchors", anchor_rep["distinct_anchors"])]
              + [(f"  {k}", v) for k, v in sorted(anchor_rep["kinds"].items(),
                                                  key=lambda kv: -kv[1])])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26

    # ---------------- data sheets ----------------
    def add_sheet(title, subset, note=""):
        sh = wb.create_sheet(safe_sheet_name(title))
        hdr_row = 1
        if note:
            sh.append([note])
            sh.cell(row=1, column=1).font = Font(italic=True, color="FF666666")
            sh.append([])
            hdr_row = 3
        sh.append([label for _, label in COLUMNS])
        for c in sh[hdr_row]:
            c.fill, c.font = header_fill, header_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
        sh.row_dimensions[hdr_row].height = 28

        vcol = [k for k, _ in COLUMNS].index("verdict") + 1
        for r in subset:
            sh.append([_val(r, k) for k, _ in COLUMNS])
            fill = VERDICT_COLORS.get(r.get("verdict"), (None, None))[0]
            if fill:
                sh.cell(row=sh.max_row, column=vcol).fill = PatternFill("solid", fgColor=fill)

        widths = {
            "URL": 52, "Domain": 26, "Verdict": 11, "Score": 7, "Tier": 6,
            "Trust tier": 26, "DA": 6, "PA": 6, "Spam Score %": 12, "HTTP": 7,
            "Status": 15, "Link live?": 10, "Followed?": 10, "rel=": 18,
            "Anchor text": 32, "noindex?": 9, "What to do": 46, "Issues found": 70,
            "Positives": 30, "Page title": 38, "Final URL": 44,
            "Redirect chain": 44, "Your notes": 24, "Reached stage": 15,
            "Why it stopped": 62, "Home page title": 34, "Home page spam": 20,
            "Placement": 13, "Sitewide?": 10, "Network cluster": 14,
            "Cluster signals": 46, "CMS / generator": 24, "Redirects to": 24,
            "Language": 10, "Lang mismatch?": 13, "IP": 16, "Reached via": 14,
            "Bad TLS cert?": 12, "No HTTPS?": 11, "Server": 20,
            "On master list?": 14, "Master status": 15, "Matched by": 12,
            "Sheet disagrees?": 24,
            "Master DA": 11, "Master SS": 11, "First recorded": 14,
            "Relevance matched": 44, "Relevance missing": 30, "Sitewide finding": 52,
            "Sitewide ratio": 13, "Pages sampled": 13, "Hosted by": 26,
            "IP range": 20, "Shared host?": 12, "Wayback snapshot": 44,
            "Snapshot date": 13, "Link in snapshot?": 15, "Old anchor text": 30,
            "Archive finding": 54, "Browser re-checked?": 16, "Browser finding": 46,
        }
        for i, (_, label) in enumerate(COLUMNS, 1):
            sh.column_dimensions[get_column_letter(i)].width = widths.get(
                label, max(11, min(len(label) + 3, 22)))

        if subset:
            last = sh.max_row
            sh.freeze_panes = sh.cell(row=hdr_row + 1, column=1)
            sh.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(COLUMNS))}{last}"
            scol = get_column_letter([k for k, _ in COLUMNS].index("score") + 1)
            sh.conditional_formatting.add(
                f"{scol}{hdr_row + 1}:{scol}{last}",
                ColorScaleRule(start_type="num", start_value=0, start_color="FFF8696B",
                               mid_type="num", mid_value=60, mid_color="FFFFEB84",
                               end_type="num", end_value=100, end_color="FF63BE7B"))
        return sh

    add_sheet("All links", rows, "Every backlink, worst first. Use the filter arrows on any column.")

    notes = {
        "TOXIC": "Real spam signals. Your disavow candidates - verify each one first.",
        "DEAD": "404/410/DNS. The page or domain is gone.",
        "LINK_LOST": "Page alive but your link was removed. Outreach, NOT a disavow.",
        "LOW_VALUE": "Harmless but passes no ranking signal. Do NOT disavow these.",
        "BLOCKED": "Returned 401/403/429/5xx or had a broken certificate. Check by hand.",
        "REVIEW": "Mixed signals - needs your judgement.",
        "GOOD": "Working links. No action needed.",
    }
    for v in VERDICT_ORDER:
        subset = [r for r in rows if r.get("verdict") == v]
        if subset:
            add_sheet(v.replace("_", " ").title(), subset, notes.get(v, ""))

    # status-based sheets, so "show me everything that 404s" is one click
    for key, label, _ in STATUS_GROUPS:
        subset = [r for r in rows if status_group(r) == key]
        if subset:
            add_sheet(f"HTTP {label[:24]}", subset,
                      f"Every link whose status is {label}.")

    highval = [r for r in rows if r.get("tier") in ("A", "B")]
    if highval:
        add_sheet("Authority (A+B)", highval,
                  ".gov/.edu/.ac and curated high-authority domains - your most valuable links.")
    sitewide = [r for r in rows if r.get("link_is_sitewide")]
    if sitewide:
        add_sheet("Sitewide links", sitewide,
                  "Footer/sidebar/blogroll links. Google devalues these heavily.")

    # ---------------- anchors ----------------
    if anchor_rep and anchor_rep["total_anchors"]:
        sh = wb.create_sheet("Anchor text")
        sh.append(["Anchor text distribution"])
        sh.cell(row=1, column=1).font = Font(bold=True, size=13)
        sh.append([])
        for wa in anchor_rep["warnings"]:
            sh.append([f"WARNING: {wa}"])
            sh.cell(row=sh.max_row, column=1).font = Font(color="FFB00020", bold=True)
        if anchor_rep["warnings"]:
            sh.append([])
        hdr = sh.max_row + 1
        sh.append(["Anchor text", "Count", "Share %", "Type"])
        for c in sh[hdr]:
            c.fill, c.font = header_fill, header_font
        total = max(1, anchor_rep["total_anchors"])
        for text, n in anchor_rep["top_anchors"]:
            sh.append([text, n, round(100 * n / total, 1),
                       anchors_mod.classify_anchor(text, anchor_rep["brand_terms"])])
        sh.column_dimensions["A"].width = 52
        for col in "BCD":
            sh.column_dimensions[col].width = 14

    # ---------------- networks ----------------
    if net and net.get("clusters"):
        sh = wb.create_sheet("Link networks")
        sh.append(["Detected link-network clusters"])
        sh.cell(row=1, column=1).font = Font(bold=True, size=13)
        sh.append(["A cluster means several audited domains share multiple footprint "
                   "signals - judge them together, not one by one."])
        sh.append([])
        sh.append(["Cluster", "Size", "Shared signals", "Domains"])
        for c in sh[4]:
            c.fill, c.font = header_fill, header_font
        for c in net["clusters"]:
            sh.append([c["id"], c["size"], "; ".join(c["signals"]), ", ".join(c["members"])])
        sh.column_dimensions["A"].width = 12
        sh.column_dimensions["B"].width = 8
        sh.column_dimensions["C"].width = 60
        sh.column_dimensions["D"].width = 90

    wb.save(path)


# ---------------------------------------------------------------------------
# HTML dashboard
# ---------------------------------------------------------------------------
# A dashboard is operated, not read. So the order is: what needs doing, then
# the shape of the profile, then the rows themselves -- and state is encoded in
# form (severity stripe, score meter, pill) as well as in number, so what needs
# attention reads at a glance instead of being counted.

# What to do about a link, in precedence order. One action per row, because a
# list where a row appears three times is a list nobody works through.
ACTIONS = [
    ("correct_sheet_bad", "Recorded clean, now toxic", "critical",
     "Your sheet says No Issues but the live check found spam. Disavow, and fix the row."),
    ("correct_sheet_clean", "Recorded spammy, now clean", "warn",
     "Your sheet says Spammy but nothing is wrong now. You may be disavowing it for nothing."),
    ("disavow", "Review and disavow", "critical",
     "Real spam signals. Verify each one, then use disavow_merged.txt."),
    ("reclaim", "Ask for the link back", "critical",
     "The page is alive but your link was removed. This is an email, not a disavow."),
    ("fix", "Fix, redirect or drop", "warn",
     "The page or domain is gone. Ask for a fix, redirect it, or drop it from the list."),
    ("upgrade", "Ask for an upgrade", "warn",
     "Decent links held back by rel=nofollow or by sitting in sitewide boilerplate."),
    ("verify", "Check by hand", "info",
     "We could not read these - 401/403/429, a 5xx, or a broken certificate."),
    ("metrics", "Get DA / PA", "info",
     "Passed every earlier gate and still has no authority data. See da_pa_queue.txt."),
]
ACTION_LABELS = {k: (label, pri, blurb) for k, label, pri, blurb in ACTIONS}


def action_group(row: dict) -> str:
    """The single next action for this link."""
    v = row.get("verdict")
    # A contradiction with your own records outranks the generic action: it is
    # both a decision about the link and a correction to make to the sheet.
    dis = master_disagreement(row)
    if dis == "clean_now_bad":
        return "correct_sheet_bad"
    if dis == "spammy_now_clean":
        return "correct_sheet_clean"
    if v == "TOXIC":
        return "disavow"
    if v == "LINK_LOST":
        return "reclaim"
    if v == "DEAD":
        return "fix"
    if v == "BLOCKED":
        return "verify"
    score = row.get("score") or 0
    if row.get("link_found") and score >= 50 and (
            row.get("is_nofollow") or row.get("is_sponsored")
            or row.get("is_ugc") or row.get("link_is_sitewide")):
        return "upgrade"
    if row.get("needs_metrics") and row.get("da") in (None, ""):
        return "metrics"
    return "none"


# Field groups for the expandable per-row detail. This is where the 79 columns
# become useful without making the table unreadable.
DETAIL_GROUPS = [
    ("Your link", [
        ("Found on page", "link_found"), ("Followed", "is_followed"),
        ("rel", "link_rel"), ("Anchor text", "anchor_texts"),
        ("Placement", "link_placement"), ("Sitewide", "sitewide_note"),
        ("Hidden block", "link_in_hidden_block"),
    ]),
    ("Request", [
        ("HTTP", "status_code"), ("Status", "status_verdict"),
        ("Reached via", "scheme_used"), ("Bad TLS cert", "tls_invalid"),
        ("HTTPS unavailable", "https_unavailable"), ("Response ms", "response_ms"),
        ("Redirect chain", "redirect_chain"), ("Redirects to", "redirect_target_domain"),
        ("Root-domain check", "root_note"), ("Error", "error"),
        ("Browser re-check", "render_note"),
    ]),
    ("Page", [
        ("Title", "page_title"), ("CMS", "generator"), ("Language", "lang"),
        ("Words", "word_count"), ("Outbound links", "outbound_links"),
        ("noindex", "is_noindex"), ("robots.txt blocked", "robots_blocked"),
        ("Relevance", "relevance_score"), ("Matched", "relevance_matched"),
        ("Missing", "relevance_missing"),
    ]),
    ("Home page", [
        ("HTTP", "home_status_code"), ("Title", "home_title"),
        ("Spam", "home_content_spam_categories"),
        ("Outbound links", "home_outbound_links"), ("Parked", "home_parked_markers"),
    ]),
    ("Spam signals", [
        ("URL", "url_spam_categories"), ("Content", "content_spam_categories"),
        ("Keywords", "content_spam_keywords"), ("Parked", "parked_markers"),
        ("Paid-link markers", "paid_link_markers"),
        ("Directory markers", "link_directory_markers"),
        ("Risky TLD", "spam_tld"), ("Free subdomain", "free_subdomain"),
    ]),
    ("Authority", [
        ("DA", "da"), ("PA", "pa"), ("Spam Score", "spam_score"),
        ("Backlinks", "backlinks"), ("Referring domains", "referring_domains"),
        ("Source", "metrics_source"), ("Tier", "tier_label"),
        ("Authority group", "authority_group"),
    ]),
    ("Hosting", [
        ("IP", "ip"), ("Hosted by", "net_org"), ("Range", "net_range"),
        ("Large shared host", "shared_host"), ("Server", "server"),
        ("Network cluster", "network_id"), ("Cluster size", "network_size"),
        ("Shared signals", "network_signals"),
    ]),
    ("Archive", [
        ("Snapshot", "archive_url"), ("Snapshot date", "archive_date"),
        ("Link in snapshot", "archive_link_found"),
        ("Old anchor", "archive_anchor"), ("Finding", "archive_note"),
    ]),
    ("Your master list", [
        ("Already recorded", "in_master"), ("Status you gave it", "master_status"),
        ("Matched by", "master_match"), ("Matched host", "master_host"),
        ("Recorded DA", "master_da"), ("Recorded Spam Score", "master_spam_score"),
        ("First recorded", "master_first_seen"), ("Source", "master_source"),
        ("Sheet disagrees", "master_disagrees"),
        ("Your notes", "master_notes"),
    ]),
    ("Pipeline", [
        ("Reached stage", "gate_stage"), ("Why", "gate_reason"),
        ("Your notes", "notes"),
    ]),
]

_DASH_CSS = """
*{box-sizing:border-box}
:root{
 --ground:#F3F6F8;--surface:#FFF;--surface-2:#EDF1F4;--surface-3:#F8FAFB;
 --line:#DCE4E8;--line-soft:#EBF0F3;
 --ink:#0F1A21;--ink-2:#3A4B55;--muted:#657983;
 --accent:#0B6E7F;--accent-soft:#E3F0F2;
 --good:#15803D;--good-soft:#E7F6EC;
 --review:#A2570A;--review-soft:#FDF3E4;
 --low:#64748B;--low-soft:#EFF2F5;
 --lost:#C2410C;--lost-soft:#FEF0E7;
 --dead:#9A3412;--dead-soft:#FBEDE7;
 --blocked:#6B7280;--blocked-soft:#F1F3F5;
 --toxic:#B01C1C;--toxic-soft:#FDECEC;
 --shadow:0 1px 2px rgba(15,26,33,.05),0 10px 26px -18px rgba(15,26,33,.22);
}
@media (prefers-color-scheme:dark){:root:not([data-theme="light"]){
 --ground:#0A1216;--surface:#111D23;--surface-2:#16242B;--surface-3:#0E1A20;
 --line:#24383F;--line-soft:#1B2A31;
 --ink:#E8F0F2;--ink-2:#B7C9CF;--muted:#8CA3AB;
 --accent:#3FB3C2;--accent-soft:#123138;
 --good:#4FBE7B;--good-soft:#11291B;
 --review:#D99442;--review-soft:#2A1F10;
 --low:#94A3B8;--low-soft:#1A242B;
 --lost:#F08A55;--lost-soft:#2B1810;
 --dead:#E08060;--dead-soft:#28150F;
 --blocked:#9AA6AF;--blocked-soft:#1B2226;
 --toxic:#E9736F;--toxic-soft:#2C1315;
 --shadow:0 1px 2px rgba(0,0,0,.4),0 12px 30px -20px rgba(0,0,0,.8);
}}
:root[data-theme="dark"]{
 --ground:#0A1216;--surface:#111D23;--surface-2:#16242B;--surface-3:#0E1A20;
 --line:#24383F;--line-soft:#1B2A31;
 --ink:#E8F0F2;--ink-2:#B7C9CF;--muted:#8CA3AB;
 --accent:#3FB3C2;--accent-soft:#123138;
 --good:#4FBE7B;--good-soft:#11291B;
 --review:#D99442;--review-soft:#2A1F10;
 --low:#94A3B8;--low-soft:#1A242B;
 --lost:#F08A55;--lost-soft:#2B1810;
 --dead:#E08060;--dead-soft:#28150F;
 --blocked:#9AA6AF;--blocked-soft:#1B2226;
 --toxic:#E9736F;--toxic-soft:#2C1315;
 --shadow:0 1px 2px rgba(0,0,0,.4),0 12px 30px -20px rgba(0,0,0,.8);
}
body{margin:0;background:var(--ground);color:var(--ink);
 font:15px/1.55 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;
 -webkit-font-smoothing:antialiased}
.wrap{max-width:1560px;margin:0 auto;padding:26px 20px 70px}
h1{margin:0;font-size:23px;letter-spacing:-.02em;font-weight:700}
h2{margin:0 0 13px;font-size:11.5px;text-transform:uppercase;letter-spacing:.07em;
 color:var(--muted);font-weight:700}
.sub{color:var(--muted);font-size:12.5px;margin-top:5px}
.mono{font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace}
.num{font-variant-numeric:tabular-nums}

/* KPI cards */
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(138px,1fr));
 gap:10px;margin:20px 0}
.kpi{background:var(--surface);border:1px solid var(--line);border-radius:11px;
 padding:14px 16px;box-shadow:var(--shadow);text-align:left;font:inherit;
 cursor:default}
.kpi.clickable{cursor:pointer}
.kpi.clickable:hover{border-color:var(--accent)}
.kpi b{display:block;font-size:26px;line-height:1.12;letter-spacing:-.02em;
 font-variant-numeric:tabular-nums}
.kpi span{display:block;color:var(--muted);font-size:11.5px;margin-top:3px}

/* action queue */
.todo{background:var(--surface);border:1px solid var(--line);border-radius:12px;
 padding:18px 20px;margin-bottom:16px;box-shadow:var(--shadow)}
.todo-list{display:grid;grid-template-columns:repeat(auto-fit,minmax(268px,1fr));gap:9px}
.todo-item{display:flex;gap:13px;align-items:flex-start;text-align:left;
 background:var(--surface-3);border:1px solid var(--line-soft);border-left:3px solid var(--low);
 border-radius:9px;padding:12px 14px;cursor:pointer;font:inherit;color:inherit;width:100%}
.todo-item:hover{border-color:var(--accent);background:var(--surface-2)}
.todo-item.critical{border-left-color:var(--toxic)}
.todo-item.warn{border-left-color:var(--review)}
.todo-item.info{border-left-color:var(--accent)}
.todo-n{font-size:24px;font-weight:700;line-height:1;min-width:38px;
 font-variant-numeric:tabular-nums;letter-spacing:-.02em}
.todo-item.critical .todo-n{color:var(--toxic)}
.todo-item.warn .todo-n{color:var(--review)}
.todo-item.info .todo-n{color:var(--accent)}
.todo-t{display:block;font-weight:600;font-size:13.5px;line-height:1.3}
.todo-d{display:block;color:var(--muted);font-size:12px;margin-top:3px;line-height:1.4}
.todo-clear{color:var(--muted);font-size:13px}

/* panels */
.panels{display:grid;grid-template-columns:repeat(auto-fit,minmax(310px,1fr));
 gap:14px;margin-bottom:16px}
.panel{background:var(--surface);border:1px solid var(--line);border-radius:12px;
 padding:17px 19px;box-shadow:var(--shadow)}
.bar{display:flex;align-items:center;gap:9px;margin:6px 0;font-size:12.5px;
 background:none;border:0;padding:2px 0;width:100%;font:inherit;color:inherit;
 text-align:left;cursor:pointer;border-radius:5px}
.bar:hover .bl{color:var(--accent)}
.bar[disabled]{cursor:default}
.bl{flex:0 0 132px;color:var(--ink-2);font-size:12.5px}
.bt{flex:1;background:var(--surface-2);border-radius:5px;height:8px;overflow:hidden}
.bt i{display:block;height:100%;border-radius:5px}
.bn{flex:0 0 40px;text-align:right;font-weight:700;font-variant-numeric:tabular-nums;
 font-size:12.5px}
.hint{color:var(--muted);font-size:11.5px;margin-top:11px;line-height:1.45}
.warnbox{background:var(--toxic-soft);border:1px solid var(--toxic);color:var(--toxic);
 border-radius:8px;padding:8px 11px;font-size:12px;margin:8px 0}
table.mini{width:100%;border-collapse:collapse;margin-top:11px;font-size:12px}
table.mini th{text-align:left;color:var(--muted);font-weight:600;padding:4px 5px;
 border-bottom:1px solid var(--line)}
table.mini td{padding:4px 5px;border-bottom:1px solid var(--line-soft)}
.cl{padding:8px 0;border-bottom:1px solid var(--line-soft);font-size:12.5px}
.cl:last-child{border-bottom:0}
.cl b{color:var(--accent)}
.dim{color:var(--muted);font-size:11.5px;margin-top:2px;line-height:1.45}

/* sticky toolbar */
.bar-wrap{position:sticky;top:0;z-index:20;padding:9px 0 11px;background:var(--ground)}
.filters{background:var(--surface);border:1px solid var(--line);border-radius:12px;
 padding:13px 15px;box-shadow:var(--shadow)}
.fgrid{display:grid;grid-template-columns:repeat(auto-fit,minmax(172px,1fr));gap:9px 12px}
.fsel{display:flex;flex-direction:column;gap:3px}
.fsel span{font-size:10px;text-transform:uppercase;letter-spacing:.06em;
 color:var(--muted);font-weight:700}
.fsel select{appearance:none;-webkit-appearance:none;width:100%;
 padding:7px 28px 7px 10px;border:1px solid var(--line);border-radius:7px;
 background:var(--surface) url("data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%23657983' d='M6 8.8 1.6 4.4l.9-.9L6 7l3.5-3.5.9.9z'/%3E%3C/svg%3E") no-repeat right 9px center;
 font-size:12.5px;color:var(--ink);cursor:pointer;font-family:inherit}
.fsel select:focus{outline:2px solid var(--accent);outline-offset:1px}
.fsel select.active{border-color:var(--accent);background-color:var(--accent-soft);font-weight:600}
.frow{display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-top:11px;
 padding-top:11px;border-top:1px solid var(--line-soft)}
#q{flex:1;min-width:180px;padding:8px 11px;border:1px solid var(--line);
 border-radius:7px;font-size:12.5px;font-family:inherit;background:var(--surface);
 color:var(--ink)}
.btn{cursor:pointer;border:1px solid var(--line);background:var(--surface);
 border-radius:7px;padding:8px 12px;font-size:12.5px;color:var(--ink-2);
 font-family:inherit;white-space:nowrap}
.btn:hover{background:var(--surface-2);border-color:var(--accent);color:var(--accent)}
#count{font-size:12px;color:var(--muted);margin-left:auto;font-variant-numeric:tabular-nums}

/* table */
.tw{background:var(--surface);border:1px solid var(--line);border-radius:12px;
 overflow:auto;box-shadow:var(--shadow)}
table.main{border-collapse:collapse;width:100%;min-width:1080px}
table.main th{position:sticky;top:0;background:var(--surface-2);text-align:left;
 padding:10px 12px;font-size:10.5px;text-transform:uppercase;letter-spacing:.05em;
 color:var(--muted);border-bottom:1px solid var(--line);z-index:5;white-space:nowrap;
 cursor:pointer;user-select:none;font-weight:700}
table.main th:hover{color:var(--accent)}
table.main th::after{content:"";font-size:9px;opacity:.5;margin-left:4px}
table.main th.asc::after{content:"\\2191";opacity:1}
table.main th.desc::after{content:"\\2193";opacity:1}
table.main td{padding:10px 12px;border-bottom:1px solid var(--line-soft);
 vertical-align:top;font-size:12.5px}
tr.row{cursor:pointer}
tr.row:hover td{background:var(--surface-3)}
tr.row.open td{background:var(--surface-2)}
td.stripe{padding:0;width:4px;border-bottom:1px solid var(--line-soft)}
tr.row.v-good td.stripe{background:var(--good)}
tr.row.v-review td.stripe{background:var(--review)}
tr.row.v-low_value td.stripe{background:var(--low)}
tr.row.v-link_lost td.stripe{background:var(--lost)}
tr.row.v-dead td.stripe{background:var(--dead)}
tr.row.v-blocked td.stripe{background:var(--blocked)}
tr.row.v-toxic td.stripe{background:var(--toxic)}
.u{max-width:330px}
.u a{color:var(--accent);text-decoration:none;word-break:break-all;font-weight:500}
.u a:hover{text-decoration:underline}
.dom{color:var(--muted);font-size:11px;margin-top:2px}
.tags{margin-top:4px;display:flex;flex-wrap:wrap;gap:3px}
.tag{background:var(--surface-2);color:var(--ink-2);border-radius:4px;
 padding:1px 5px;font-size:10px;font-weight:600}
.tag.warn{background:var(--toxic-soft);color:var(--toxic)}
.tag.net{background:var(--accent-soft);color:var(--accent)}
.tag.ok{background:var(--good-soft);color:var(--good)}
.pill{padding:2px 8px;border-radius:999px;font-size:10.5px;font-weight:700;
 white-space:nowrap;display:inline-block}
.p-good{background:var(--good-soft);color:var(--good)}
.p-review{background:var(--review-soft);color:var(--review)}
.p-low_value{background:var(--low-soft);color:var(--low)}
.p-link_lost{background:var(--lost-soft);color:var(--lost)}
.p-dead{background:var(--dead-soft);color:var(--dead)}
.p-blocked{background:var(--blocked-soft);color:var(--blocked)}
.p-toxic{background:var(--toxic-soft);color:var(--toxic)}
.sc{display:flex;align-items:center;gap:7px;min-width:74px}
.sc b{font-variant-numeric:tabular-nums;font-size:13px;min-width:22px;text-align:right}
.meter{flex:1;height:5px;background:var(--surface-2);border-radius:4px;overflow:hidden}
.meter i{display:block;height:100%;border-radius:4px}
.act{max-width:210px;color:var(--ink-2)}
.iss{max-width:360px;color:var(--muted);font-size:11.5px;line-height:1.45}
.ctr{text-align:center;color:var(--muted)}

/* row detail */
tr.detail{display:none}
tr.detail.show{display:table-row}
tr.detail td{background:var(--surface-3);padding:0;border-bottom:2px solid var(--line)}
.dwrap{padding:16px 18px;display:grid;
 grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:16px}
.dgroup h4{margin:0 0 7px;font-size:10.5px;text-transform:uppercase;
 letter-spacing:.06em;color:var(--accent);font-weight:700}
.dgroup dl{margin:0;display:grid;grid-template-columns:auto 1fr;gap:3px 10px;
 font-size:12px}
.dgroup dt{color:var(--muted);white-space:nowrap}
.dgroup dd{margin:0;color:var(--ink-2);word-break:break-word}
.dfull{grid-column:1/-1;border-top:1px solid var(--line-soft);padding-top:12px}
.dfull .lab{font-size:10.5px;text-transform:uppercase;letter-spacing:.06em;
 color:var(--muted);font-weight:700;margin-bottom:4px}
.dfull p{margin:0 0 9px;font-size:12.5px;color:var(--ink-2);line-height:1.5}

.note{background:var(--accent-soft);border:1px solid var(--accent);border-radius:10px;
 padding:12px 15px;font-size:12.5px;color:var(--ink-2);margin-bottom:16px}
.note.warnnote{background:var(--review-soft);border-color:var(--review)}
footer{margin-top:20px;color:var(--muted);font-size:11.5px;text-align:center}
:focus-visible{outline:2px solid var(--accent);outline-offset:2px}
@media (max-width:700px){.bl{flex-basis:104px}#count{margin-left:0;width:100%}}
@media (prefers-reduced-motion:reduce){*{transition:none!important;animation:none!important}}
"""

_DASH_JS = """
(function(){
 var tb=document.getElementById('tb'),
     rows=[].slice.call(tb.querySelectorAll('tr.row')),
     selects=[].slice.call(document.querySelectorAll('.fsel select')),
     q=document.getElementById('q'),
     countEl=document.getElementById('count'),
     emptyRow=document.getElementById('empty');

 function detailOf(r){ return r.nextElementSibling; }

 function visibleRows(){
   return rows.filter(function(r){ return r.style.display!=='none'; });
 }

 function apply(){
   var term=(q.value||'').toLowerCase(), shown=0, active=0;
   selects.forEach(function(sl){
     var on=sl.value!=='ALL';
     sl.classList.toggle('active',on);
     if(on) active++;
   });
   rows.forEach(function(r){
     var ok=true;
     for(var i=0;i<selects.length;i++){
       var sl=selects[i];
       if(sl.value!=='ALL' && r.getAttribute('data-'+sl.getAttribute('data-g'))!==sl.value){
         ok=false; break;
       }
     }
     if(ok && term) ok = r.textContent.toLowerCase().indexOf(term)>-1;
     r.style.display = ok ? '' : 'none';
     var d=detailOf(r);
     if(d && d.classList.contains('detail') && !ok) d.classList.remove('show');
     if(ok) shown++;
   });
   if(emptyRow) emptyRow.style.display = shown ? 'none' : '';
   countEl.textContent = shown+' of '+rows.length+' shown'
     + (active||term ? ' \\u00b7 '+(active+(term?1:0))+' filter(s) active' : '');
 }

 function setFilters(obj){
   selects.forEach(function(sl){
     var g=sl.getAttribute('data-g');
     sl.value = (obj && obj[g]) ? obj[g] : 'ALL';
   });
   q.value='';
   apply();
   var tw=document.querySelector('.tw');
   if(tw) tw.scrollIntoView({behavior:'smooth',block:'start'});
 }

 selects.forEach(function(sl){ sl.addEventListener('change',apply); });
 q.addEventListener('input',apply);

 document.getElementById('reset').addEventListener('click',function(){ setFilters(null); });

 [].slice.call(document.querySelectorAll('[data-set]')).forEach(function(el){
   el.addEventListener('click',function(){
     var v={}; try{ v=JSON.parse(el.getAttribute('data-set')); }catch(e){}
     setFilters(v);
   });
 });

 /* expand a row for the full findings */
 rows.forEach(function(r){
   r.addEventListener('click',function(ev){
     if(ev.target.closest('a')) return;
     var d=detailOf(r);
     if(!d||!d.classList.contains('detail')) return;
     var open=d.classList.toggle('show');
     r.classList.toggle('open',open);
   });
 });

 /* copy + csv of whatever is currently on screen */
 function flash(btn,msg){
   var old=btn.textContent; btn.textContent=msg;
   setTimeout(function(){ btn.textContent=old; },1400);
 }
 var copyBtn=document.getElementById('copyurls');
 copyBtn.addEventListener('click',function(){
   var urls=visibleRows().map(function(r){ return r.getAttribute('data-url'); }).join('\\n');
   if(!urls){ flash(copyBtn,'nothing to copy'); return; }
   if(navigator.clipboard && navigator.clipboard.writeText){
     navigator.clipboard.writeText(urls).then(function(){
       flash(copyBtn,'copied '+visibleRows().length);
     },function(){ flash(copyBtn,'copy blocked'); });
   } else {
     var ta=document.createElement('textarea');
     ta.value=urls; document.body.appendChild(ta); ta.select();
     try{ document.execCommand('copy'); flash(copyBtn,'copied '+visibleRows().length); }
     catch(e){ flash(copyBtn,'copy blocked'); }
     document.body.removeChild(ta);
   }
 });

 var csvBtn=document.getElementById('csvbtn');
 csvBtn.addEventListener('click',function(){
   var vis=visibleRows();
   if(!vis.length){ flash(csvBtn,'nothing to export'); return; }
   var head=['URL','Domain','Verdict','Score','DA','SpamScore','HTTP','Action','Issues'];
   var lines=[head.join(',')];
   vis.forEach(function(r){
     var c=r.getAttribute('data-csv')||'';
     lines.push(c);
   });
   var blob=new Blob([lines.join('\\n')],{type:'text/csv;charset=utf-8'});
   var a=document.createElement('a');
   a.href=URL.createObjectURL(blob);
   a.download='filtered_backlinks.csv';
   document.body.appendChild(a); a.click(); document.body.removeChild(a);
   flash(csvBtn,'exported '+vis.length);
 });

 /* sorting */
 var ths=[].slice.call(document.querySelectorAll('table.main th')), dir={};
 ths.forEach(function(th,i){
   th.addEventListener('click',function(){
     var kind=th.getAttribute('data-s')||'text';
     dir[i]=!dir[i];
     var mul=dir[i]?1:-1;
     ths.forEach(function(x){ x.classList.remove('asc','desc'); });
     th.classList.add(dir[i]?'asc':'desc');
     var pairs=rows.map(function(r){ return [r,detailOf(r)]; });
     pairs.sort(function(a,b){
       var x=a[0].cells[i]?a[0].cells[i].innerText.trim():'',
           y=b[0].cells[i]?b[0].cells[i].innerText.trim():'';
       if(kind==='num'){
         var nx=parseFloat(x.replace(/[^0-9.\\-]/g,'')), ny=parseFloat(y.replace(/[^0-9.\\-]/g,''));
         if(isNaN(nx))nx=-1; if(isNaN(ny))ny=-1;
         return (nx-ny)*mul;
       }
       return x.localeCompare(y)*mul;
     });
     pairs.forEach(function(p){
       tb.appendChild(p[0]);
       if(p[1]&&p[1].classList.contains('detail')) tb.appendChild(p[1]);
     });
     if(emptyRow) tb.appendChild(emptyRow);
   });
 });

 /* keyboard: / focuses search, Esc clears everything */
 document.addEventListener('keydown',function(e){
   if(e.key==='/' && document.activeElement!==q){ e.preventDefault(); q.focus(); q.select(); }
   else if(e.key==='Escape'){ setFilters(null); q.blur(); }
 });

 apply();
})();
"""


def master_disagreement(row: dict) -> str:
    """
    Does the live check contradict what your master list says?

    This is a finding in its own right, and the reason a known domain is still
    fully re-checked every run. A real case: anibookmark.com was recorded as
    "No Issues" while its live pages carried Parimatch, Mostbet and Bet7k
    Casino links. Without surfacing the contradiction you would keep trusting
    a stale row and never know which sheet entries had rotted.

    Both directions matter. "was clean, now toxic" is a domain to disavow AND a
    row to correct. "was spammy, now fine" is a link you may be needlessly
    disavowing.

    Returns "" | "clean_now_bad" | "spammy_now_clean" | "clean_now_dead"
    """
    if not row.get("in_master"):
        return ""
    ms = (row.get("master_status") or "").lower()
    was_clean = "no issue" in ms
    was_spammy = "spam" in ms
    verdict = row.get("verdict")

    if was_clean and verdict == "TOXIC":
        return "clean_now_bad"
    if was_clean and verdict in ("DEAD",):
        return "clean_now_dead"
    if was_spammy and verdict in ("GOOD", "REVIEW"):
        return "spammy_now_clean"
    return ""


DISAGREEMENT_TEXT = {
    "clean_now_bad": ("Recorded clean, now toxic",
                      "Your sheet says No Issues but the live check found real spam "
                      "signals. Disavow it, and correct the row."),
    "clean_now_dead": ("Recorded clean, now dead",
                       "Your sheet says No Issues but the domain no longer resolves "
                       "or the page is gone."),
    "spammy_now_clean": ("Recorded spammy, now clean",
                         "Your sheet says Spammy but the live check found nothing "
                         "wrong. You may be disavowing this needlessly - check it."),
}


def _known_key(row: dict) -> str:
    """Bucket a row by what your master list already says about it."""
    if not row.get("in_master"):
        return "new"
    ms = (row.get("master_status") or "").lower()
    if "spam" in ms:
        return "spammy"
    if "no issue" in ms:
        return "clean"
    return "other"


def _fmt(v):
    """Render a field value for the detail panel, or '' to omit it."""
    if v is None or v == "" or v == [] or v == {}:
        return ""
    if isinstance(v, bool):
        return "yes" if v else ""
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def write_html(rows: list, path: str, summary: dict, meta: dict,
               anchor_rep: dict | None = None, net: dict | None = None):
    e = lambda s: _html.escape(str(s if s is not None else ""))
    total = max(1, summary["total"])
    for r in rows:
        dis = master_disagreement(r)
        r["master_disagrees"] = DISAGREEMENT_TEXT.get(dis, ("", ""))[0] if dis else ""
        r.setdefault("action_group", action_group(r))

    vcolor = {"GOOD": "var(--good)", "REVIEW": "var(--review)",
              "LOW_VALUE": "var(--low)", "LINK_LOST": "var(--lost)",
              "DEAD": "var(--dead)", "BLOCKED": "var(--blocked)",
              "TOXIC": "var(--toxic)"}

    # ---------------- KPI cards ----------------
    live_pct = round(100 * summary["links_live"] / max(1, summary["links_checked"]))
    kpis = [
        ("Backlinks", summary["total"], "var(--ink)", None),
        ("Average score", summary["avg_score"], "var(--ink)", None),
        ("Working for you", summary["verdicts"].get("GOOD", 0), "var(--good)",
         {"verdict": "GOOD"}),
        ("Need action", sum(summary["verdicts"].get(v, 0)
                            for v in ("TOXIC", "LINK_LOST", "DEAD")), "var(--toxic)", None),
        ("Links still live", f"{live_pct}%" if summary["links_checked"] else "n/a",
         "var(--accent)", None),
        ("Reached DA/PA stage", summary.get("gate_passed", 0), "var(--accent)",
         {"stage": "4-metrics"}),
        ("New to you", summary.get("brand_new", 0), "var(--accent)", {"known": "new"}),
    ]
    kpi_html = "".join(
        (f'<button class="kpi clickable" data-set=\'{json.dumps(f)}\'>'
         if f else '<div class="kpi">')
        + f'<b style="color:{c}">{e(v)}</b><span>{e(k)}</span>'
        + ("</button>" if f else "</div>")
        for k, v, c, f in kpis)

    # ---------------- action queue ----------------
    counts = Counter(r["action_group"] for r in rows)
    items = []
    for key, label, pri, blurb in ACTIONS:
        n = counts.get(key, 0)
        if not n:
            continue
        items.append(
            f'<button class="todo-item {pri}" data-set=\'{{"action":"{key}"}}\'>'
            f'<span class="todo-n">{n}</span><span>'
            f'<span class="todo-t">{e(label)}</span>'
            f'<span class="todo-d">{e(blurb)}</span></span></button>')
    if net and net.get("clusters"):
        nc = len(net["clusters"])
        items.append(
            f'<button class="todo-item critical" data-set=\'{{"network":"yes"}}\'>'
            f'<span class="todo-n">{nc}</span><span>'
            f'<span class="todo-t">Link network{"s" if nc != 1 else ""} detected</span>'
            f'<span class="todo-d">{net["domains_in_clusters"]} domains share multiple '
            f'footprint signals. Judge each cluster as one decision.</span></span></button>')
    todo_html = ("".join(items) if items else
                 '<div class="todo-clear">Nothing needs action. Every link is either '
                 'working or harmlessly low-value.</div>')

    # ---------------- panels ----------------
    def bars(entries):
        out = []
        for label, count, color, filt in entries:
            if not count:
                continue
            attr = (f' data-set=\'{json.dumps(filt)}\'' if filt else " disabled")
            out.append(
                f'<button class="bar"{attr}><span class="bl">{e(label)}</span>'
                f'<span class="bt"><i style="width:{count / total * 100:.1f}%;'
                f'background:{color}"></i></span>'
                f'<span class="bn">{count}</span></button>')
        return "".join(out) or '<div class="hint">none</div>'

    verdict_bars = bars([(v.replace("_", " ").title(), summary["verdicts"].get(v, 0),
                          vcolor.get(v, "var(--accent)"), {"verdict": v})
                         for v in VERDICT_ORDER])
    status_bars = bars([(label, summary["status_groups"].get(key, 0), col, {"status": key})
                        for key, label, col in STATUS_GROUPS])
    tier_bars = bars([(TIER_LABELS[k], summary["tiers"].get(k, 0),
                       "var(--accent)", {"tier": k}) for k in "ABCD"])
    stage_bars = bars([
        ("Stage 1 - not live", summary["stages"].get("1-live", 0),
         "var(--toxic)", {"stage": "1-live"}),
        ("Stage 2 - page", summary["stages"].get("2-page-content", 0),
         "var(--lost)", {"stage": "2-page-content"}),
        ("Stage 3 - home page", summary["stages"].get("3-homepage", 0),
         "var(--review)", {"stage": "3-homepage"}),
        ("Stage 4 - DA/PA", summary["stages"].get("4-metrics", 0),
         "var(--good)", {"stage": "4-metrics"}),
    ])
    verify_bars = bars([
        ("Link still present", summary["links_live"], "var(--good)", None),
        ("Followed", summary["links_followed"], "var(--accent)", None),
        ("In-content", summary.get("links_incontent", 0), "var(--accent)",
         {"placement": "in-content"}),
        ("Sitewide footer", summary.get("links_sitewide", 0), "var(--review)",
         {"placement": "boilerplate"}),
        ("Link missing", summary["links_checked"] - summary["links_live"],
         "var(--lost)", {"placement": "none"}),
        ("noindex pages", summary["noindex"], "var(--toxic)", None),
    ])

    anchor_panel = ""
    if anchor_rep and anchor_rep["total_anchors"]:
        tot = max(1, anchor_rep["total_anchors"])
        kc = {"branded": "var(--good)", "naked-url": "var(--accent)",
              "keyword": "var(--review)", "generic": "var(--low)",
              "image-empty": "var(--toxic)"}
        kb = "".join(
            f'<button class="bar" disabled><span class="bl">{e(k)}</span>'
            f'<span class="bt"><i style="width:{n / tot * 100:.1f}%;'
            f'background:{kc.get(k, "var(--accent)")}"></i></span>'
            f'<span class="bn">{n}</span></button>'
            for k, n in sorted(anchor_rep["kinds"].items(), key=lambda kv: -kv[1]))
        top = "".join(
            f"<tr><td>{e(t)}</td><td class='bn num'>{n}</td>"
            f"<td class='bn num'>{100 * n / tot:.0f}%</td></tr>"
            for t, n in anchor_rep["top_anchors"][:6])
        warn = "".join(f'<div class="warnbox">{e(w)}</div>'
                       for w in anchor_rep["warnings"])
        anchor_panel = (
            f'<div class="panel"><h2>Anchor text &middot; {anchor_rep["total_anchors"]} '
            f'anchors, {anchor_rep["distinct_anchors"]} distinct</h2>{kb}{warn}'
            f'<table class="mini"><thead><tr><th>Most common</th><th>n</th>'
            f'<th>%</th></tr></thead><tbody>{top}</tbody></table></div>')

    master_panel = ""
    if summary.get("in_master") or summary.get("brand_new"):
        mb = bars([
            ("Already spammy", summary.get("master_spammy", 0), "var(--toxic)",
             {"known": "spammy"}),
            ("Already cleared", summary.get("master_clean", 0), "var(--good)",
             {"known": "clean"}),
            ("Recorded, no status",
             summary.get("in_master", 0) - summary.get("master_spammy", 0)
             - summary.get("master_clean", 0), "var(--low)", {"known": "other"}),
            ("New to you", summary.get("brand_new", 0), "var(--accent)",
             {"known": "new"}),
            ("Sheet now wrong", summary.get("master_disagrees", 0), "var(--toxic)",
             {"action": "correct_sheet_bad"}),
        ])
        master_panel = (
            f'<div class="panel"><h2>Your master list &middot; '
            f'{summary.get("in_master", 0)} of {summary["total"]} already recorded</h2>{mb}'
            f'<div class="hint">Rows tagged <b>existing data</b> are domains you have '
            f'already ruled on. They are still fully re-checked every run - only the '
            f'DA/Spam Score is reused, and only while it is recent. <b>New to you</b> is '
            f'where fresh judgement is actually needed.'
            + (f' <b>{summary.get("master_disagrees", 0)} row(s) now contradict your '
               f'sheet</b> - that is what re-checking known domains is for.'
               if summary.get("master_disagrees") else "")
            + '</div></div>')

    net_panel = ""
    if net and net.get("clusters"):
        cl = "".join(
            f'<div class="cl"><b>{e(c["id"])}</b> &middot; {c["size"]} domains'
            f'<div class="dim">{e("; ".join(c["signals"]))}</div>'
            f'<div class="dim mono">{e(", ".join(c["members"][:6]))}'
            f'{"&hellip;" if len(c["members"]) > 6 else ""}</div></div>'
            for c in net["clusters"][:5])
        net_panel = (
            f'<div class="panel"><h2>Link networks &middot; {len(net["clusters"])} '
            f'cluster(s)</h2>{cl}'
            f'<div class="hint">A cluster means several domains share more than one '
            f'footprint signal. Shared hosting alone is discounted, so these are worth '
            f'judging together rather than one at a time.</div></div>')

    # ---------------- notice: what could not be checked ----------------
    # Worth saying loudly rather than leaving as an absence. Someone reading a
    # report full of LOW_VALUE rows should know whether that is a judgement or
    # simply a check that never ran.
    n_unverifiable = sum(1 for r in rows if not r.get("target_checked"))
    skipped_note = ""
    if n_unverifiable:
        pct = round(100 * n_unverifiable / total)
        all_of_them = n_unverifiable == summary["total"]
        skipped_note = (
            f'<div class="note warnnote"><b>Link verification did not run on '
            f'{n_unverifiable} of {summary["total"]} rows ({pct}%).</b> '
            + ("Every row was supplied as a bare domain, so there is no linking page "
               if all_of_them else
               "Those rows had no linking page to check, so ")
            + "to look at. That means <b>&ldquo;is my link still there?&rdquo;, its anchor "
              "text, and whether it sits in an article or a footer are all unanswered</b> "
              "&mdash; the three most valuable checks here. Re-run with referring "
              "<i>page</i> URLs (Ahrefs and Search Console both export a "
              "&ldquo;Referring page&rdquo; column) to get them.</div>")

    # ---------------- filter dropdowns ----------------
    def build_select(group, label, entries, all_label):
        opts = [f'<option value="ALL">{e(all_label)} ({summary["total"]})</option>']
        for key, text, count in entries:
            if not count:
                continue
            opts.append(f'<option value="{e(key)}">{e(text)} ({count})</option>')
        return (f'<label class="fsel"><span>{e(label)}</span>'
                f'<select data-g="{group}">{"".join(opts)}</select></label>')

    placement_counts = Counter(
        "none" if not r.get("link_found") else (r.get("link_placement") or "unknown")
        for r in rows)
    stage_counts = Counter(r.get("gate_stage", "") for r in rows)

    sels = "".join([
        build_select("action", "Next action",
                     [(k, ACTION_LABELS[k][0], counts.get(k, 0)) for k, _, _, _ in ACTIONS]
                     + [("none", "No action needed", counts.get("none", 0))],
                     "All actions"),
        build_select("verdict", "Verdict",
                     [(v, v.replace("_", " ").title(), summary["verdicts"].get(v, 0))
                      for v in VERDICT_ORDER], "All verdicts"),
        build_select("status", "HTTP status",
                     [(k, lab, summary["status_groups"].get(k, 0))
                      for k, lab, _ in STATUS_GROUPS], "All statuses"),
        build_select("tier", "Trust tier",
                     [(k, TIER_LABELS[k], summary["tiers"].get(k, 0)) for k in "ABCD"],
                     "All tiers"),
        build_select("stage", "Pipeline stage",
                     [(k, {"1-live": "Stage 1 - not live",
                           "2-page-content": "Stage 2 - page content",
                           "3-homepage": "Stage 3 - home page",
                           "4-metrics": "Stage 4 - reached DA/PA"}.get(k, k), n)
                      for k, n in sorted(stage_counts.items()) if k], "All stages"),
        build_select("placement", "Link placement", [
            ("in-content", "In-content (editorial)", placement_counts.get("in-content", 0)),
            ("boilerplate", "Sitewide boilerplate", placement_counts.get("boilerplate", 0)),
            ("unknown", "Position unclear", placement_counts.get("unknown", 0)),
            ("none", "No link found", placement_counts.get("none", 0)),
        ], "All placements"),
        build_select("network", "Link network", [
            ("yes", "In a detected cluster", summary.get("in_network", 0)),
            ("no", "Not in a cluster", summary["total"] - summary.get("in_network", 0)),
        ], "All domains"),
        build_select("known", "Master list", [
            ("spammy", "Existing data - spammy", summary.get("master_spammy", 0)),
            ("clean", "Existing data - no issues", summary.get("master_clean", 0)),
            ("other", "Existing data - no status",
             summary.get("in_master", 0) - summary.get("master_spammy", 0)
             - summary.get("master_clean", 0)),
            ("new", "New domain", summary.get("brand_new", 0)),
        ], "Known and new"),
    ])

    # ---------------- table ----------------
    trs = []
    for idx, r in enumerate(rows):
        v = r.get("verdict", "")
        vk = v.lower()
        score = r.get("score") or 0
        da = r.get("da")
        ss = r.get("spam_score")
        ag = r.get("action_group", "none")

        tags = []
        if r.get("network_id"):
            tags.append(f'<span class="tag net">{e(r["network_id"])}</span>')
        if r.get("link_is_sitewide"):
            tags.append('<span class="tag">sitewide</span>')
        elif r.get("link_placement") == "in-content":
            tags.append('<span class="tag ok">in-content</span>')
        if r.get("is_noindex"):
            tags.append('<span class="tag warn">noindex</span>')
        if r.get("tls_invalid"):
            tags.append('<span class="tag warn">bad cert</span>')
        if r.get("lang_mismatch"):
            tags.append(f'<span class="tag">lang {e(r.get("lang", ""))}</span>')
        if r.get("rendered") and r.get("link_found"):
            tags.append('<span class="tag ok">JS link</span>')
        if r.get("master_disagrees"):
            tags.append(f'<span class="tag warn" title="{e(DISAGREEMENT_TEXT[master_disagreement(r)][1])}">'
                        f'sheet out of date</span>')
        if r.get("in_master"):
            ms = (r.get("master_status") or "").strip()
            cls = ("warn" if "spam" in ms.lower()
                   else ("ok" if "no issue" in ms.lower() else ""))
            label = f"existing data{f' &middot; {e(ms)}' if ms else ''}"
            tags.append(f'<span class="tag {cls}" title="Already on your master list'
                        f'{f" as {e(ms)}" if ms else ""}, matched by '
                        f'{e(r.get("master_match", ""))}">{label}</span>')
        else:
            tags.append('<span class="tag">new domain</span>')
        if r.get("archive_link_found"):
            tags.append('<span class="tag ok">archived</span>')
        if r.get("sitewide_sampled", 0) >= 2 and r.get("sitewide_ratio", 0) >= 1.0:
            tags.append('<span class="tag warn">on every page</span>')

        link_cell = ("&#10003; live" if r.get("link_found")
                     else ("&#10007; gone" if r.get("target_checked") else "&ndash;"))
        sub = []
        if r.get("link_found"):
            sub.append("followed" if r.get("is_followed")
                       else e(r.get("link_rel") or "nofollow"))

        csv_line = ",".join('"' + str(x).replace('"', '""') + '"' for x in [
            r.get("url", ""), r.get("registered", ""), v, score,
            "" if da in (None, "") else da, "" if ss in (None, "") else ss,
            r.get("status_code") or "", ACTION_LABELS.get(ag, ("", "", ""))[0],
            (r.get("issues") or "")[:300]])

        trs.append(
            f'<tr class="row v-{vk}" data-verdict="{e(v)}" '
            f'data-status="{e(status_group(r))}" data-tier="{e(r.get("tier", ""))}" '
            f'data-stage="{e(r.get("gate_stage", ""))}" data-action="{e(ag)}" '
            f'data-placement="{e("none" if not r.get("link_found") else (r.get("link_placement") or "unknown"))}" '
            f'data-network="{"yes" if r.get("network_id") else "no"}" '
            f'data-known="{_known_key(r)}" '
            f'data-url="{e(r.get("url", ""))}" data-csv="{e(csv_line)}">'
            f'<td class="stripe"></td>'
            f'<td class="u"><a href="{e(r.get("url", ""))}" target="_blank" '
            f'rel="noopener noreferrer">{e(r.get("url", "")[:88])}</a>'
            f'<div class="dom">{e(r.get("registered", ""))} &middot; '
            f'{e(r.get("tier_label", ""))}</div>'
            f'<div class="tags">{"".join(tags)}</div></td>'
            f'<td><span class="pill p-{vk}">{e(v.replace("_", " "))}</span></td>'
            f'<td><span class="sc"><b>{score}</b><span class="meter">'
            f'<i style="width:{max(2, min(100, score))}%;'
            f'background:{vcolor.get(v, "var(--accent)")}"></i></span></span></td>'
            f'<td class="bn num">{e("" if da in (None, "") else int(float(da)))}</td>'
            f'<td class="bn num">{e("" if ss in (None, "") else str(int(float(ss))) + "%")}</td>'
            f'<td class="bn num">{e(r.get("status_code") or "err")}</td>'
            f'<td>{link_cell}<div class="dom">{" &middot; ".join(sub)}</div></td>'
            f'<td class="act">{e(r.get("action", ""))}</td>'
            f'<td class="iss">{e((r.get("issues") or "")[:190])}</td>'
            "</tr>")

        # ---- expandable detail row ----
        groups = []
        for gname, fields in DETAIL_GROUPS:
            pairs = [(lab, _fmt(r.get(key))) for lab, key in fields]
            pairs = [(lab, val) for lab, val in pairs if val]
            if not pairs:
                continue
            dl = "".join(f"<dt>{e(lab)}</dt><dd>{e(val[:300])}</dd>" for lab, val in pairs)
            groups.append(f'<div class="dgroup"><h4>{e(gname)}</h4><dl>{dl}</dl></div>')
        full = ""
        if r.get("issues") and r["issues"] != "none found":
            full += f'<div class="lab">Every issue found</div><p>{e(r["issues"])}</p>'
        if r.get("wins"):
            full += f'<div class="lab">Positives</div><p>{e(r["wins"])}</p>'
        if full:
            groups.append(f'<div class="dfull">{full}</div>')
        trs.append(
            f'<tr class="detail"><td colspan="10"><div class="dwrap">'
            f'{"".join(groups) or "<div class=dgroup>No further detail.</div>"}'
            f"</div></td></tr>")

    doc = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Backlink Audit - {e(meta.get('target', ''))}</title>
<style>{_DASH_CSS}</style></head><body><div class="wrap">

<header>
  <h1>Backlink Audit</h1>
  <div class="sub">Target <b>{e(meta.get('target', ''))}</b> &middot;
  {summary['total']} backlinks &middot; run {e(meta.get('when', ''))} &middot;
  DA source: {e(meta.get('metrics_provider', 'none'))}</div>
</header>

<div class="kpis">{kpi_html}</div>

<div class="todo">
  <h2>What needs doing</h2>
  <div class="todo-list">{todo_html}</div>
  <div class="hint">Each box filters the table below. One action per link, so
  nothing appears twice.</div>
</div>

{skipped_note}
<div class="note"><b>Before you disavow anything.</b>
<b>Toxic</b> means real spam signals. <b>Low value</b> means harmless but passes
nothing &mdash; do <u>not</u> disavow those, you would only hurt yourself.
<b>Link lost</b> is an email to send. <b>Blocked</b> means we could not read the
page, so it is not judged. Click any row for the full findings.</div>

<div class="panels">
  <div class="panel"><h2>Verdicts</h2>{verdict_bars}</div>
  <div class="panel"><h2>HTTP status</h2>{status_bars}</div>
  <div class="panel"><h2>Staged pipeline</h2>{stage_bars}
    <div class="hint">A link that fails a stage never reaches the next one,
    which is why only {summary.get('gate_passed', 0)} of {summary['total']}
    needed an authority lookup.</div></div>
  <div class="panel"><h2>Domain trust tiers</h2>{tier_bars}</div>
  <div class="panel"><h2>Link verification</h2>{verify_bars}</div>
  {master_panel}{anchor_panel}{net_panel}
</div>

<div class="bar-wrap"><div class="filters">
  <div class="fgrid">{sels}</div>
  <div class="frow">
    <input id="q" type="search" placeholder="Search URL, domain, issue, anchor, cluster&hellip;  ( / )">
    <button class="btn" id="reset" type="button">Reset</button>
    <button class="btn" id="copyurls" type="button">Copy URLs</button>
    <button class="btn" id="csvbtn" type="button">Export CSV</button>
    <span id="count"></span>
  </div>
</div></div>

<div class="tw"><table class="main"><thead><tr>
<th style="width:4px"></th>
<th data-s="text">URL / domain</th><th data-s="text">Verdict</th>
<th data-s="num">Score</th><th data-s="num">DA</th><th data-s="num">SS</th>
<th data-s="num">HTTP</th><th data-s="text">Your link</th>
<th data-s="text">What to do</th><th data-s="text">Issues found</th>
</tr></thead><tbody id="tb">{''.join(trs)}
<tr id="empty" style="display:none"><td colspan="10" class="ctr"
 style="padding:34px">Nothing matches these filters. Press
 <b>Esc</b> or Reset to clear them.</td></tr>
</tbody></table></div>

<footer>Generated by SEO Backlink Audit &middot; scores are heuristic &mdash;
open a page before you disavow it &middot; <b>/</b> to search, <b>Esc</b> to reset</footer>
</div><script>{_DASH_JS}</script></body></html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


def write_json(rows: list, path: str, summary: dict, meta: dict,
               anchor_rep: dict | None = None, net: dict | None = None):
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "summary": summary, "anchors": anchor_rep,
                   "networks": net, "rows": rows}, f, indent=1, default=str)
