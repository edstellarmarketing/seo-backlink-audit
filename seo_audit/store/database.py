"""
The project's database: your master domain list, plus audit history.

Why SQLite rather than more CSVs
--------------------------------
The master sheet is 10,443 domains. Answering "have I seen this domain before,
and what did I decide about it?" against a CSV means re-parsing the whole file
on every run and matching by hand. Answering it against an indexed table is a
single lookup, and the answer can then feed three separate things:

  1. an "existing data" tag in the report, so you instantly see which links you
     have already ruled on;
  2. DA and Spam Score you ALREADY OWN, which removes those domains from the
     paste-and-import queue entirely;
  3. a place for each run's results to accumulate, so change over time is a
     query rather than a folder of JSON files.

One file, no server, no dependency: sqlite3 ships with Python.

Tables
------
  master_domains   your sheet: host, DA, SS, status, where it came from
  audit_runs       one row per run
  audit_links      one row per link per run  (history and trends)

Matching is done on BOTH the exact host and the registered domain, because a
sheet listing `blog.example.com` should still flag a link on
`example.com/page`, and vice versa - just with different confidence, which the
match_kind column records.
"""

import csv
import json
import os
import re
import sqlite3
import threading
from datetime import datetime, timezone

from seo_audit.analysis import domains as dom

SCHEMA = """
CREATE TABLE IF NOT EXISTS master_domains (
    host              TEXT PRIMARY KEY,
    registered        TEXT NOT NULL,
    da                REAL,
    spam_score        REAL,
    status            TEXT,
    source            TEXT,
    notes             TEXT,
    first_seen        TEXT NOT NULL,
    last_seen         TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS ix_master_registered ON master_domains(registered);
CREATE INDEX IF NOT EXISTS ix_master_status     ON master_domains(status);

CREATE TABLE IF NOT EXISTS audit_runs (
    run_id      TEXT PRIMARY KEY,
    started_at  TEXT NOT NULL,
    target      TEXT,
    link_count  INTEGER,
    summary     TEXT
);

CREATE TABLE IF NOT EXISTS audit_links (
    run_id      TEXT NOT NULL,
    url         TEXT NOT NULL,
    host        TEXT,
    registered  TEXT,
    verdict     TEXT,
    score       REAL,
    da          REAL,
    spam_score  REAL,
    status_code TEXT,
    tier        TEXT,
    link_found  INTEGER,
    is_followed INTEGER,
    action      TEXT,
    issues      TEXT,
    PRIMARY KEY (run_id, url)
);
CREATE INDEX IF NOT EXISTS ix_links_registered ON audit_links(registered);
CREATE INDEX IF NOT EXISTS ix_links_verdict    ON audit_links(verdict);

CREATE TABLE IF NOT EXISTS meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);
"""

# Column-name detection for imports, so any export of the sheet works.
_COL = {
    "host":       re.compile(r"^\s*(domain|domains|host|url|website|site|referring\s*domain)\s*$", re.I),
    "da":         re.compile(r"^\s*(da|domain\s*authority|domain_authority)\s*$", re.I),
    "spam_score": re.compile(r"^\s*(ss|spam\s*score|spam_score|spamscore|spam)\s*%?\s*$", re.I),
    "status":     re.compile(r"^\s*(status|verdict|decision|result|action)\s*$", re.I),
    "notes":      re.compile(r"^\s*(notes?|comment|remark|reason)\s*$", re.I),
}
_NUM = re.compile(r"-?\d+(?:\.\d+)?")


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s or s.upper() in ("N/A", "NA", "-", "--", "NULL", "NONE"):
        return None
    m = _NUM.search(s)
    return float(m.group()) if m else None


def _now():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _clean_host(value: str) -> str:
    """A sheet cell to a bare hostname. Tolerates full URLs and stray spaces."""
    s = str(value or "").strip().strip('"').strip("'").lower()
    if not s or " " in s:
        return ""
    host = dom.host_of(s)
    return host.lstrip(".") if host else ""


class AuditDb:
    def __init__(self, path: str):
        self.path = path
        self._local = threading.local()
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with self._connect() as c:
            c.executescript(SCHEMA)

    # sqlite connections are not shareable across threads, so give each
    # thread its own. The audit runs in a pool; lookups happen from workers.
    def _connect(self):
        conn = getattr(self._local, "conn", None)
        if conn is None:
            conn = sqlite3.connect(self.path, timeout=30)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            self._local.conn = conn
        return conn

    # ------------------------------------------------------------------
    # import
    # ------------------------------------------------------------------
    def import_sheet(self, path: str, source: str | None = None,
                     default_status: str | None = None) -> dict:
        """
        Load a master sheet (.csv / .xlsx) into master_domains.

        Existing hosts are UPDATED, not duplicated, and first_seen is preserved
        so you can still tell how long a domain has been on your list.
        """
        source = source or os.path.basename(path)
        rows = self._read_rows(path)
        if not rows:
            return {"file": source, "read": 0, "imported": 0, "updated": 0,
                     "skipped": 0, "error": "no rows found"}

        header_idx, colmap = self._find_header(rows)
        if colmap is None:
            return {"file": source, "read": len(rows), "imported": 0, "updated": 0,
                     "skipped": len(rows), "error": "no recognisable Domain column"}

        now = _now()
        seen, imported, updated, skipped = set(), 0, 0, 0
        conn = self._connect()
        with conn:
            for raw in rows[header_idx + 1:]:
                def cell(key):
                    j = colmap.get(key)
                    return raw[j] if j is not None and len(raw) > j else ""

                host = _clean_host(cell("host"))
                if not host or host in seen:
                    skipped += 1
                    continue
                seen.add(host)
                reg = dom.registered_domain(host) or host
                status = (str(cell("status")).strip() or default_status or "").strip()
                rec = (host, reg, _num(cell("da")), _num(cell("spam_score")),
                       status, source, str(cell("notes")).strip(), now, now)

                cur = conn.execute("SELECT host FROM master_domains WHERE host=?", (host,))
                if cur.fetchone():
                    conn.execute(
                        """UPDATE master_domains
                              SET registered=?, da=COALESCE(?,da),
                                  spam_score=COALESCE(?,spam_score),
                                  status=CASE WHEN ?<>'' THEN ? ELSE status END,
                                  source=?, notes=CASE WHEN ?<>'' THEN ? ELSE notes END,
                                  last_seen=?
                            WHERE host=?""",
                        (reg, rec[2], rec[3], status, status, source,
                         rec[6], rec[6], now, host))
                    updated += 1
                else:
                    conn.execute(
                        """INSERT INTO master_domains
                           (host,registered,da,spam_score,status,source,notes,
                            first_seen,last_seen)
                           VALUES (?,?,?,?,?,?,?,?,?)""", rec)
                    imported += 1
            conn.execute("INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
                         ("last_import", json.dumps({"file": source, "at": now})))
        return {"file": source, "read": len(rows) - header_idx - 1,
                "imported": imported, "updated": updated, "skipped": skipped,
                "error": ""}

    @staticmethod
    def _read_rows(path: str) -> list:
        ext = os.path.splitext(path)[1].lower()
        if ext in (".xlsx", ".xlsm"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return []
            wb = load_workbook(path, read_only=True, data_only=True)
            out = []
            for ws in wb.worksheets:
                out.extend(list(r) for r in ws.iter_rows(values_only=True))
            wb.close()
            return out
        with open(path, newline="", encoding="utf-8-sig", errors="ignore") as f:
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            return [list(r) for r in csv.reader(f, dialect)]

    @staticmethod
    def _find_header(rows: list):
        """
        Locate the header row and map our field names onto column indexes.

        The real sheet has TWO blocks of Domain/DA/SS/Status columns side by
        side with the second one empty, so we take the FIRST match for each
        field and ignore the duplicates rather than getting confused by them.
        """
        for idx, row in enumerate(rows[:25]):
            cells = [str(c or "").strip() for c in row]
            found = {}
            for j, c in enumerate(cells):
                for key, pat in _COL.items():
                    if key not in found and pat.match(c):
                        found[key] = j
            if "host" in found:
                return idx, found
        # No header at all: assume the first column is the domain.
        if rows and _clean_host(rows[0][0] if rows[0] else ""):
            return -1, {"host": 0}
        return 0, None

    # ------------------------------------------------------------------
    # lookup
    # ------------------------------------------------------------------
    EMPTY = {"in_master": False, "master_status": "", "master_da": None,
             "master_spam_score": None, "master_match": "", "master_host": "",
             "master_source": "", "master_first_seen": "", "master_notes": ""}

    def lookup(self, url: str) -> dict:
        """
        Is this URL's domain already on the master list?

        Tries the exact host first, then the registered domain. The match_kind
        matters: an exact-host hit is a decision you made about THIS site; a
        registered-domain hit is a decision about the parent domain, which is
        strong but not identical.
        """
        host = dom.host_of(url)
        if not host:
            return dict(self.EMPTY)
        reg = dom.registered_domain(url)
        conn = self._connect()

        row = conn.execute("SELECT * FROM master_domains WHERE host=?", (host,)).fetchone()
        kind = "host"
        if row is None and reg and reg != host:
            row = conn.execute("SELECT * FROM master_domains WHERE host=?", (reg,)).fetchone()
            kind = "registered"
        if row is None and reg:
            row = conn.execute(
                "SELECT * FROM master_domains WHERE registered=? ORDER BY LENGTH(host) LIMIT 1",
                (reg,)).fetchone()
            kind = "registered"
        if row is None:
            return dict(self.EMPTY)
        return {
            "in_master": True,
            "master_status": row["status"] or "",
            "master_da": row["da"],
            "master_spam_score": row["spam_score"],
            "master_match": kind,
            "master_host": row["host"],
            "master_source": row["source"] or "",
            "master_first_seen": (row["first_seen"] or "")[:10],
            "master_notes": row["notes"] or "",
        }

    def metrics_for(self, registered_domains: list, max_age_days: float = 30.0) -> dict:
        """
        DA / Spam Score you already own, for the metrics layer.

        `max_age_days` is the point of this method. A DA figure recorded last
        week is worth reusing; one from two years ago is not, and silently
        reusing it would quietly rot the whole audit. Anything older than the
        limit is simply not returned, so it flows on to the DA/PA queue for a
        refresh instead.

        Note what this does NOT affect: the live checks. Status, HTTPS, the
        link itself, the page content and the home page are always re-checked,
        for every domain, however well known it is. Only the authority numbers
        are ever reused.

        Returns {registered_domain: {da, spam_score, source, age_days}}.
        """
        out = {}
        if not registered_domains:
            return out
        conn = self._connect()
        cutoff = ""
        if max_age_days and max_age_days > 0:
            from datetime import timedelta
            cutoff = (datetime.now(timezone.utc)
                      - timedelta(days=float(max_age_days))).isoformat(timespec="seconds")
        CHUNK = 400
        for i in range(0, len(registered_domains), CHUNK):
            chunk = [d for d in registered_domains[i:i + CHUNK] if d]
            if not chunk:
                continue
            marks = ",".join("?" * len(chunk))
            sql = (f"""SELECT registered, da, spam_score, status, last_seen
                         FROM master_domains
                        WHERE registered IN ({marks})
                          AND (da IS NOT NULL OR spam_score IS NOT NULL)""")
            args = list(chunk)
            if cutoff:
                sql += " AND last_seen >= ?"
                args.append(cutoff)
            sql += " ORDER BY LENGTH(host)"
            for row in conn.execute(sql, args):
                out.setdefault(row["registered"], {
                    "da": row["da"], "spam_score": row["spam_score"],
                    "source": "master-sheet", "recorded": (row["last_seen"] or "")[:10],
                })
        return out

    def stale_metrics(self, registered_domains: list, max_age_days: float = 30.0) -> list:
        """
        Domains on the master list whose DA is too old to reuse. These are the
        ones worth re-checking, as opposed to ones you have never checked.
        """
        if not registered_domains or not max_age_days or max_age_days <= 0:
            return []
        from datetime import timedelta
        cutoff = (datetime.now(timezone.utc)
                  - timedelta(days=float(max_age_days))).isoformat(timespec="seconds")
        conn = self._connect()
        out = []
        CHUNK = 400
        for i in range(0, len(registered_domains), CHUNK):
            chunk = [d for d in registered_domains[i:i + CHUNK] if d]
            if not chunk:
                continue
            marks = ",".join("?" * len(chunk))
            for row in conn.execute(
                    f"""SELECT DISTINCT registered FROM master_domains
                         WHERE registered IN ({marks}) AND last_seen < ?""",
                    chunk + [cutoff]):
                out.append(row["registered"])
        return out

    # ------------------------------------------------------------------
    # run history
    # ------------------------------------------------------------------
    def record_run(self, run_id: str, target: str, rows: list, summary: dict):
        conn = self._connect()
        with conn:
            conn.execute(
                """INSERT OR REPLACE INTO audit_runs
                   (run_id,started_at,target,link_count,summary)
                   VALUES (?,?,?,?,?)""",
                (run_id, _now(), target, len(rows), json.dumps(summary, default=str)))
            conn.executemany(
                """INSERT OR REPLACE INTO audit_links
                   (run_id,url,host,registered,verdict,score,da,spam_score,
                    status_code,tier,link_found,is_followed,action,issues)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [(run_id, r.get("url", ""), r.get("host", ""), r.get("registered", ""),
                  r.get("verdict", ""), r.get("score"), r.get("da"), r.get("spam_score"),
                  str(r.get("status_code") or ""), r.get("tier", ""),
                  1 if r.get("link_found") else 0, 1 if r.get("is_followed") else 0,
                  r.get("action_group", ""), (r.get("issues") or "")[:900])
                 for r in rows])

    def add_from_audit(self, rows: list, only_verdicts=("TOXIC",)) -> int:
        """
        Fold this run's findings back into the master list, so a domain the
        audit judges toxic is remembered next time. Never overwrites a status
        you set by hand from the sheet -- it only fills a blank one.
        """
        now = _now()
        added = 0
        conn = self._connect()
        with conn:
            for r in rows:
                if r.get("verdict") not in only_verdicts:
                    continue
                host = r.get("host") or dom.host_of(r.get("url", ""))
                if not host:
                    continue
                reg = r.get("registered") or dom.registered_domain(host) or host
                exists = conn.execute("SELECT host, status FROM master_domains WHERE host=?",
                                      (host,)).fetchone()
                if exists:
                    if not (exists["status"] or "").strip():
                        conn.execute(
                            "UPDATE master_domains SET status=?, last_seen=? WHERE host=?",
                            ("Spammy", now, host))
                    continue
                conn.execute(
                    """INSERT INTO master_domains
                       (host,registered,da,spam_score,status,source,notes,
                        first_seen,last_seen)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (host, reg, r.get("da"), r.get("spam_score"), "Spammy",
                     "audit", (r.get("issues") or "")[:400], now, now))
                added += 1
        return added

    # ------------------------------------------------------------------
    def stats(self) -> dict:
        conn = self._connect()
        q = lambda sql, *a: conn.execute(sql, a).fetchone()[0]
        by_status = {r["status"] or "(blank)": r["n"] for r in conn.execute(
            "SELECT status, COUNT(*) n FROM master_domains GROUP BY status ORDER BY n DESC")}
        return {
            "master_domains": q("SELECT COUNT(*) FROM master_domains"),
            "with_da": q("SELECT COUNT(*) FROM master_domains WHERE da IS NOT NULL"),
            "with_spam_score": q("SELECT COUNT(*) FROM master_domains WHERE spam_score IS NOT NULL"),
            "by_status": by_status,
            "runs": q("SELECT COUNT(*) FROM audit_runs"),
            "links_recorded": q("SELECT COUNT(*) FROM audit_links"),
            # WAL mode keeps recent writes in a sidecar file, so the main file
            # alone under-reports the size (it read 4 KB for 10k rows).
            "db_bytes": sum(os.path.getsize(self.path + suf)
                            for suf in ("", "-wal", "-shm")
                            if os.path.exists(self.path + suf)),
        }

    def close(self):
        conn = getattr(self._local, "conn", None)
        if conn is not None:
            conn.close()
            self._local.conn = None


def open_db(root: str, cfg: dict | None = None) -> AuditDb:
    cfg = cfg or {}
    rel = ((cfg.get("database") or {}).get("path") or "data/audit.db")
    return AuditDb(os.path.join(root, rel))
